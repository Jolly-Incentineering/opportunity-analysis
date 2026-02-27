"""
cheatsheet_gen.py — Generate company and campaign cheat sheets as a combined PDF.

Usage:
    python .claude/scripts/cheatsheet_gen.py --company "Company Name"

Output:
    Clients/[Company]/4. Reports/Cheat Sheets/[Company] Cheat Sheet.pdf

Requires (one of):
    pip install weasyprint        # primary renderer
    pip install playwright && playwright install chromium  # fallback renderer
"""
from __future__ import annotations

import sys
import json
import glob
import argparse
import os
import re
from pathlib import Path
from datetime import date

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import weasyprint
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Brand colors
# ---------------------------------------------------------------------------

NAVY    = "#123769"
NAVY_LT = "#1a4a8a"
GOLD    = "#E8A838"
GOLD_LT = "#FEF6E4"
BG      = "#F6F8FA"
WHITE   = "#FFFFFF"
GRAY_LT = "#EEF1F5"
GRAY    = "#D1D5DC"
BORDER  = "#DDE2EA"
TEXT    = "#1a1a2e"
MUTED   = "#666D80"
GREEN   = "#1a7a4a"
GREEN_LT= "#e8f6ee"
RED     = "#b03030"
RED_LT  = "#fdeaea"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def slugify(s: str) -> str:
    return re.sub(r"[^a-z0-9_]", "", s.lower().replace(" ", "_").replace("-", "_"))


def esc(s) -> str:
    return (str(s or "")
            .replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def fmt(v, prefix="$") -> str:
    """Format numeric value. Use prefix='' for plain numbers."""
    if v is None:
        return "N/A"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v)
    sign = "-" if v < 0 else ""
    av = abs(v)
    if 0 < av < 1:
        return f"{v * 100:.1f}%"
    if av == 0:
        return f"{prefix}0" if prefix else "0"
    if av >= 1_000_000_000:
        return f"{sign}{prefix}{av / 1_000_000_000:.1f}B"
    if av >= 1_000_000:
        return f"{sign}{prefix}{av / 1_000_000:.1f}M"
    if av >= 1_000:
        return f"{sign}{prefix}{av / 1_000:.1f}K"
    if v == int(v):
        return f"{sign}{prefix}{int(av)}"
    return f"{sign}{prefix}{av:.2f}" if prefix else str(round(av, 2))


def fmt_plain(v) -> str:
    return fmt(v, prefix="")


def _deep_get(d: dict, path: str, default=None):
    """Traverse nested dict/list via dot-separated path."""
    parts = path.split(".")
    cur = d
    for p in parts:
        if isinstance(cur, dict):
            cur = cur.get(p)
        elif isinstance(cur, list) and p.isdigit():
            idx = int(p)
            cur = cur[idx] if idx < len(cur) else None
        else:
            return default
        if cur is None:
            return default
    return cur


def _section(icon: str, title: str, content: str) -> str:
    """Wrap content in a standard labelled section block."""
    return (
        f'<div class="section">'
        f'<div class="section-label"><span class="s-icon">{esc(icon)}</span>{esc(title)}</div>'
        f'{content}'
        f'</div>'
    )


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------

def get_workspace_config() -> dict:
    candidates = [".claude/data/workspace_config.json"]
    jolly_ws = os.environ.get("JOLLY_WORKSPACE", "").strip().strip("\r")
    if jolly_ws:
        candidates.append(os.path.join(jolly_ws, ".claude/data/workspace_config.json"))
    for p in candidates:
        if os.path.exists(p):
            try:
                with open(p, encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return {}


def find_research_json(company: str, base_path: str = None) -> dict:
    slug = slugify(company)
    if base_path:
        client_glob = f"{base_path}/**/research_output_*.json"
    else:
        cfg = get_workspace_config()
        client_root = cfg.get("client_root", "Clients")
        client_glob = f"{client_root}/{company}/**/research_output_*.json"

    candidates = []
    for m in glob.glob(client_glob, recursive=True):
        try:
            candidates.append((os.path.getmtime(m), m))
        except OSError:
            pass
    for _, path in sorted(candidates, reverse=True):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            print(f"  Research JSON: {path}")
            return data
        except (json.JSONDecodeError, OSError):
            continue
    for p in [
        f".claude/data/research_output_{slug}.json",
        f".claude/data/research_output_{company.lower().replace(' ', '-')}.json",
    ]:
        if os.path.exists(p):
            print(f"  Research JSON (legacy): {p}")
            with open(p, encoding="utf-8") as f:
                return json.load(f)
    raise FileNotFoundError(f"No research JSON found for '{company}'")


def find_model(company: str, base_path: str = None) -> str:
    if base_path:
        model_glob = f"{base_path}/1. Model/*.xlsx"
        search_path = f"{base_path}/1. Model/"
    else:
        cfg = get_workspace_config()
        client_root = cfg.get("client_root", "Clients")
        model_glob = f"{client_root}/{company}/1. Model/*.xlsx"
        search_path = f"{client_root}/{company}/1. Model/"
    matches = glob.glob(model_glob)
    if not matches:
        raise FileNotFoundError(f"No Excel model in {search_path}")
    return sorted(matches)[-1]


def _fmt_assumption(val) -> str:
    """Format a scenario assumption value for display."""
    if val is None:
        return ""
    if isinstance(val, float):
        if 0 < val < 1:
            return f"{val:.0%}"
        if val >= 1000:
            return f"${val:,.0f}"
        if val >= 100:
            return f"{val:,.0f}"
        if val == int(val):
            return f"{int(val):,}"
        return f"{val:.2f}"
    if isinstance(val, int):
        return f"${val:,}" if val >= 1000 else f"{val:,}"
    return str(val)


def read_model_assumptions(model_path: str) -> dict:
    """Read only the scenario assumptions section from the Excel Inputs sheet.

    Returns dict: { "assumptions__{campaign_slug}": [(label, base, upside, downside), ...] }
    Does NOT read ROPS or EBITDA — those come from research JSON campaign_details.
    """
    if not OPENPYXL_AVAILABLE:
        return {}
    try:
        wb = load_workbook(model_path, data_only=True)
    except PermissionError:
        import shutil, tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            shutil.copy2(model_path, tmp_path)
            wb = load_workbook(tmp_path, data_only=True)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    if "Inputs" not in wb.sheetnames:
        return {}

    ws = wb["Inputs"]
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Find "SCENARIO ASSUMPTIONS" header row
    scenario_start = None
    for i, row in enumerate(all_rows):
        if row[1] and "SCENARIO ASSUMPTIONS" in str(row[1]):
            scenario_start = i + 1
            break
    if scenario_start is None:
        return {}

    result = {}
    current_name = None
    current_rows = []

    for row in all_rows[scenario_start:]:
        label    = row[1] if len(row) > 1 else None
        base_val = row[2] if len(row) > 2 else None
        up_val   = row[3] if len(row) > 3 else None
        dn_val   = row[4] if len(row) > 4 else None
        label_str = str(label).strip() if label else ""

        if not label_str:
            if current_name and current_rows:
                result[f"assumptions__{slugify(current_name)}"] = current_rows
                current_name = None
                current_rows = []
            continue

        if re.match(r"Campaign \d+:", label_str) and base_val is None:
            if current_name and current_rows:
                result[f"assumptions__{slugify(current_name)}"] = current_rows
            current_name = label_str.split(":", 1)[1].strip()
            current_rows = []
        elif base_val is not None:
            current_rows.append((
                label_str,
                _fmt_assumption(base_val),
                _fmt_assumption(up_val),
                _fmt_assumption(dn_val),
            ))

    if current_name and current_rows:
        result[f"assumptions__{slugify(current_name)}"] = current_rows

    return result


# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

CSS = f"""
@page {{ size: Letter; margin: 0; }}
html, body {{ width: 816px; overflow: hidden; box-sizing: border-box; }}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}

body {{
    font-family: -apple-system, 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    font-size: 10px;
    color: {TEXT};
    background: {BG};
    line-height: 1.5;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
}}

/* ── Banner ── */
.banner {{
    background: linear-gradient(135deg, {NAVY} 0%, {NAVY_LT} 100%);
    border-bottom: 3px solid {GOLD};
    padding: 10px 28px 9px;
    display: flex;
    align-items: center;
    justify-content: space-between;
}}
.banner-eyebrow {{
    font-size: 6.5px; font-weight: 700; letter-spacing: 1.8px;
    text-transform: uppercase; color: {GOLD}; margin-bottom: 2px;
}}
.banner-title {{
    font-size: 14px; font-weight: 800; color: {WHITE};
    letter-spacing: -0.3px; line-height: 1; margin-bottom: 2px;
}}
.banner-sub {{ font-size: 7.5px; color: rgba(255,255,255,0.5); }}

/* ── Body ── */
.body {{ padding: 8px 28px 20px; background: {BG}; }}

/* ── Section ── */
.section {{ margin: 10px 0 5px; break-inside: avoid; }}
.section-label + * {{ break-before: avoid; }}
.section-label {{
    font-size: 7.5px; font-weight: 800; letter-spacing: 1.4px;
    text-transform: uppercase; color: {NAVY};
    padding-bottom: 4px; border-bottom: 2px solid {GOLD};
    display: flex; align-items: center; gap: 5px;
}}
.s-icon {{
    width: 14px; height: 14px; background: {GOLD}; border-radius: 3px;
    display: inline-flex; align-items: center; justify-content: center;
    font-size: 8px; color: {NAVY}; font-weight: 900; flex-shrink: 0;
}}

/* ── KPI strip ── */
.kpi-strip {{ display: flex; gap: 0; margin-top: 6px; }}
.kpi-card {{
    flex: 1; padding: 8px 10px; background: {WHITE};
    border: 1px solid {GRAY_LT}; border-right: none; text-align: center;
}}
.kpi-card:last-child {{ border-right: 1px solid {GRAY_LT}; border-radius: 0 5px 5px 0; }}
.kpi-card:first-child {{ border-radius: 5px 0 0 5px; }}
.kpi-v {{ font-size: 18px; font-weight: 900; color: {NAVY}; letter-spacing: -0.5px; line-height: 1; }}
.kpi-l {{ font-size: 7px; font-weight: 700; color: {MUTED}; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 2px; }}
.kpi-src {{ font-size: 6.5px; color: {MUTED}; margin-top: 1px; font-style: italic; }}

/* ── Breakdown grid ── */
.breakdown-grid {{ display: flex; gap: 6px; flex-wrap: wrap; margin-top: 5px; }}
.bk-cell {{
    background: {WHITE}; border: 1px solid {GRAY_LT}; border-radius: 5px;
    padding: 6px 10px; text-align: center; min-width: 80px;
}}
.bk-v {{ font-size: 14px; font-weight: 800; color: {NAVY}; line-height: 1; }}
.bk-l {{ font-size: 7px; color: {MUTED}; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 2px; }}

/* ── Tables (shared base) ── */
.data-table, .bench-table, .tech-table, .assump-table, .objection-table {{
    width: 100%; border-collapse: collapse; margin-top: 5px; font-size: 9px;
}}
.data-table td, .bench-table td, .tech-table td, .assump-table td, .objection-table td {{
    padding: 5px 9px; border-bottom: 1px solid {GRAY_LT}; vertical-align: top;
}}
.data-table tr:last-child td, .bench-table tr:last-child td,
.tech-table tr:last-child td, .assump-table tr:last-child td,
.objection-table tr:last-child td {{ border-bottom: none; }}
.data-table tr:nth-child(even) td, .bench-table tr:nth-child(even) td,
.tech-table tr:nth-child(even) td {{ background: {GRAY_LT}; }}
.data-table td:first-child, .tech-table td:first-child {{ color: {MUTED}; width: 38%; font-weight: 600; }}

/* Table header */
.assump-hdr th, .objection-hdr td {{
    background: {NAVY}; color: {WHITE}; font-size: 7.5px;
    font-weight: 700; text-transform: uppercase; padding: 5px 9px;
}}
.bench-table td.mid {{ font-weight: 700; color: {NAVY}; }}
.col-base {{ font-weight: 600; }}
.col-up {{ color: {GREEN}; }}
.col-dn {{ color: {RED}; }}

/* ── Opportunity banner ── */
.opp-banner {{
    display: flex; gap: 0; margin: 6px 0 12px;
    border: 1px solid {GOLD}; border-radius: 7px;
    overflow: hidden; background: {GOLD_LT};
}}
.opp-item {{ flex: 1; padding: 10px 12px 9px; text-align: center; border-right: 1px solid rgba(232,168,56,0.35); }}
.opp-item:last-child {{ border-right: none; }}
.opp-v {{ font-size: 20px; font-weight: 900; color: {NAVY}; letter-spacing: -0.5px; line-height: 1; }}
.opp-l {{ font-size: 7px; font-weight: 700; color: {MUTED}; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 3px; }}

/* ── Bullets / lists ── */
.bullets {{ padding-left: 14px; font-size: 9px; margin-top: 4px; }}
.bullets li {{ margin-bottom: 3px; }}
.numbered-list {{ padding-left: 16px; font-size: 9px; margin-top: 4px; }}
.numbered-list li {{ margin-bottom: 3px; }}

/* ── Quote ── */
.quote {{
    background: {GOLD_LT}; border-left: 3px solid {GOLD};
    border-radius: 0 5px 5px 0; padding: 7px 11px;
    font-size: 9px; font-style: italic; margin-bottom: 5px; color: {TEXT};
}}
.quote-src {{ font-size: 7.5px; color: {MUTED}; margin-top: 3px; font-style: normal; }}

/* ── Pills ── */
.pills {{ display: flex; gap: 5px; flex-wrap: wrap; margin-top: 5px; }}
.pill {{
    background: {NAVY}; color: {WHITE}; border-radius: 20px;
    padding: 2px 8px; font-size: 7.5px; font-weight: 700;
    display: inline-flex; align-items: center; gap: 4px;
}}
.pill-zero {{ background: {GRAY}; color: {MUTED}; }}
.pill-val {{ font-weight: 400; opacity: 0.85; }}

/* ── Contacts ── */
.contacts-box {{
    background: {WHITE}; border: 1px solid {GRAY_LT};
    border-radius: 6px; padding: 4px 10px; margin-top: 6px;
}}
.champion-row {{ padding: 5px 0; border-bottom: 1px solid {GRAY_LT}; }}
.champion-row:last-child {{ border-bottom: none; }}
.champ-name {{ font-weight: 700; font-size: 9.5px; }}
.champ-title {{ font-size: 8.5px; color: {MUTED}; }}
.champ-note {{ font-size: 8px; color: {MUTED}; font-style: italic; }}

/* ── Deal context ── */
.deal-box {{
    background: {WHITE}; border: 1px solid {GRAY_LT};
    border-radius: 6px; padding: 7px 12px; margin-top: 5px; font-size: 9px;
}}
.deal-row {{ display: flex; gap: 8px; margin-bottom: 3px; }}
.deal-row:last-child {{ margin-bottom: 0; }}
.deal-label {{ color: {MUTED}; font-weight: 600; min-width: 90px; font-size: 8.5px; }}

/* ── Campaign cards ── */
.campaign {{
    background: {WHITE}; border: 1px solid {GRAY_LT};
    border-radius: 8px; margin-bottom: 10px;
    overflow: hidden; break-inside: avoid;
}}
.c-head {{
    background: {BG}; padding: 8px 12px 7px;
    border-bottom: 1px solid {GRAY_LT};
    display: flex; align-items: flex-start; justify-content: space-between;
}}
.c-head.high {{ background: linear-gradient(135deg, {NAVY} 0%, {NAVY_LT} 100%); }}
.c-head.high .c-title {{ color: {WHITE}; }}
.c-title {{ font-size: 11px; font-weight: 800; color: {NAVY}; letter-spacing: -0.2px; }}
.c-badges {{ display: flex; gap: 4px; align-items: center; flex-shrink: 0; }}
.badge {{
    font-size: 7px; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.5px; padding: 2px 7px; border-radius: 20px;
}}
.b-high {{ background: {GOLD}; color: {NAVY}; }}
.b-std {{ background: {GRAY_LT}; color: {MUTED}; }}
.b-interest {{ background: {GREEN_LT}; color: {GREEN}; border: 1px solid {GREEN}; }}
.c-body {{ padding: 10px 12px; }}
.c-stats {{ display: flex; gap: 0; margin-bottom: 8px; }}
.c-stat-item {{
    flex: 1; text-align: center; padding: 6px 8px;
    border: 1px solid {GRAY_LT}; border-right: none; background: {BG};
}}
.c-stat-item:first-child {{ border-radius: 5px 0 0 5px; }}
.c-stat-item:last-child {{ border-right: 1px solid {GRAY_LT}; border-radius: 0 5px 5px 0; }}
.c-stat-v {{ font-size: 14px; font-weight: 900; color: {NAVY}; line-height: 1; }}
.c-stat-l {{ font-size: 6.5px; color: {MUTED}; text-transform: uppercase; letter-spacing: 0.5px; margin-top: 2px; }}
.evidence {{
    font-size: 8.5px; font-style: italic; color: {MUTED};
    border-left: 2px solid {GOLD}; padding: 3px 8px;
    margin-bottom: 7px; background: {GOLD_LT}; border-radius: 0 3px 3px 0;
}}
.assumptions {{ margin-top: 6px; }}
.assumptions-label {{
    font-size: 7px; font-weight: 700; letter-spacing: 1px;
    text-transform: uppercase; color: {MUTED}; margin-bottom: 3px;
}}

/* ── Page break ── */
.page-break {{ break-after: page; }}

/* ── Footer ── */
.footer {{
    margin-top: 16px; padding-top: 8px;
    border-top: 1px solid {GRAY}; font-size: 7.5px; color: {MUTED};
    display: flex; justify-content: space-between;
}}
"""


# ---------------------------------------------------------------------------
# Company Cheat Sheet
# ---------------------------------------------------------------------------

def build_company_html(company: str, research: dict) -> str:
    today  = date.today().strftime("%B %d, %Y")
    basics = research.get("company_basics") or {}
    attio  = research.get("attio_insights") or {}
    comps  = research.get("comps_benchmarks") or {}
    sources= research.get("source_summary") or {}

    # ── KPI strip ──
    revenue    = basics.get("annual_revenue")
    units      = basics.get("unit_count")
    employees  = basics.get("employee_count")
    unit_label = basics.get("unit_count_label") or "Locations"

    def _src_tag(field):
        s = basics.get(f"{field}_source") or ""
        return f'<div class="kpi-src">{esc(s)}</div>' if s else ""

    kpi_content = (
        f'<div class="kpi-strip">'
        f'<div class="kpi-card"><div class="kpi-v">{fmt(revenue)}</div>'
        f'<div class="kpi-l">Annual Revenue</div>{_src_tag("annual_revenue")}</div>'
        f'<div class="kpi-card"><div class="kpi-v">{fmt_plain(units)}</div>'
        f'<div class="kpi-l">{esc(unit_label)}</div>{_src_tag("unit_count")}</div>'
        f'<div class="kpi-card"><div class="kpi-v">{fmt_plain(employees)}</div>'
        f'<div class="kpi-l">Employees</div>{_src_tag("employee_count")}</div>'
        f'</div>'
    )
    kpi_html = _section("▲", "Key Metrics", kpi_content)

    # ── Company Profile ──
    geo = basics.get("geography") or {}
    profile_rows = []
    if research.get("industry"):
        profile_rows.append(("Industry", str(research["industry"])))
    if attio.get("founded"):
        profile_rows.append(("Founded", str(attio["founded"])))
    if geo.get("hq"):
        profile_rows.append(("HQ", geo["hq"]))
    if geo.get("rank"):
        profile_rows.append(("Market Position", geo["rank"]))
    if geo.get("states"):
        profile_rows.append(("States", ", ".join(geo["states"])))
    elif geo.get("state_count"):
        profile_rows.append(("States Operated", str(geo["state_count"])))
    for key, label in [
        ("categories",   "Categories"),
        ("employee_range", "Employee Range"),
        ("estimated_arr_usd", "Est. ARR"),
        ("last_interaction", "Last Interaction"),
        ("strongest_connection_strength", "Connection"),
    ]:
        v = attio.get(key)
        if v:
            if key == "estimated_arr_usd":
                display = fmt(v)
            elif isinstance(v, list):
                display = ", ".join(str(x) for x in v)
            else:
                display = str(v)
            profile_rows.append((label, display))
    slack_list = research.get("slack_insights") or []
    slack = slack_list[0] if slack_list else {}
    for key, label in [("deal_stage", "Deal Stage"), ("primary_contact", "Primary Contact")]:
        if slack.get(key):
            profile_rows.append((label, str(slack[key])))

    profile_html = ""
    if profile_rows:
        rows_html = "".join(f"<tr><td>{esc(k)}</td><td>{esc(v)}</td></tr>" for k, v in profile_rows)
        profile_html = _section("◈", "Company Profile", f'<table class="data-table">{rows_html}</table>')

    # ── Employee Breakdown ──
    breakdown_html = ""
    bd = basics.get("employee_breakdown") or {}
    if bd:
        cells = "".join(
            f'<div class="bk-cell"><div class="bk-v">{fmt_plain(v)}</div>'
            f'<div class="bk-l">{esc(k.replace("_", " ").title())}</div></div>'
            for k, v in bd.items() if v is not None
        )
        if cells:
            breakdown_html = _section("⊞", "Employee Breakdown", f'<div class="breakdown-grid">{cells}</div>')

    # ── Industry Benchmarks ──
    bench_html = ""
    bench_rows = []
    for key, label in [
        ("ebitda_margin_pct", "EBITDA Margin"),
        ("gross_margin_pct",  "Gross Margin"),
        ("turnover_rate",     "Turnover Rate"),
        ("net_margin_pct",    "Net Margin"),
        ("labor_cost_pct",    "Labor Cost %"),
    ]:
        raw = comps.get(key)
        if raw is None:
            continue
        if isinstance(raw, dict):
            bench_rows.append((label, raw.get("low"), raw.get("mid"), raw.get("high")))
        else:
            bench_rows.append((label, None, raw, None))

    if bench_rows:
        has_range = any(lo is not None or hi is not None for _, lo, _, hi in bench_rows)
        if has_range:
            hdr = '<tr class="assump-hdr"><th>Benchmark</th><th>Low</th><th>Mid</th><th>High</th></tr>'
            rows_html = hdr + "".join(
                f'<tr><td>{esc(lbl)}</td>'
                f'<td>{fmt_plain(lo) if lo is not None else "—"}</td>'
                f'<td class="mid">{fmt_plain(mid) if mid is not None else "—"}</td>'
                f'<td>{fmt_plain(hi) if hi is not None else "—"}</td></tr>'
                for lbl, lo, mid, hi in bench_rows
            )
            bench_html = _section("≈", "Industry Benchmarks", f'<table class="bench-table">{rows_html}</table>')
        else:
            rows_html = "".join(
                f'<tr><td>{esc(lbl)}</td><td>{fmt_plain(mid) if mid is not None else "—"}</td></tr>'
                for lbl, _, mid, _ in bench_rows
            )
            bench_html = _section("≈", "Industry Benchmarks", f'<table class="data-table">{rows_html}</table>')

    # ── Research Sources ──
    pill_data = [
        ("Gong",  sources.get("gong_calls_found") or 0),
        ("Attio", sources.get("attio_records") or 0),
        ("Slack", sources.get("slack_messages") or 0),
        ("SEC",   1 if sources.get("sec_filings") else 0),
        ("Web",   sources.get("web_operations_used") or 0),
    ]
    pills = "".join(
        f'<span class="pill {"pill-zero" if not n else ""}">{esc(label)}'
        f' <span class="pill-val">{n}</span></span>'
        for label, n in pill_data
    )
    source_html = _section("◎", "Research Sources", f'<div class="pills">{pills}</div>')

    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>{CSS}</style></head>'
        f'<body>'
        f'<div class="banner"><div>'
        f'<div class="banner-eyebrow">JOLLY.COM &middot; CONFIDENTIAL</div>'
        f'<div class="banner-title">{esc(company)}</div>'
        f'<div class="banner-sub">Company Cheat Sheet &middot; {today}</div>'
        f'</div></div>'
        f'<div class="body">'
        f'{kpi_html}{profile_html}{breakdown_html}{bench_html}{source_html}'
        f'<div class="footer">'
        f'<span>Jolly.com</span><span>Confidential — Internal Use Only</span><span>{today}</span>'
        f'</div>'
        f'</div></body></html>'
    )


# ---------------------------------------------------------------------------
# Meeting Prep Page
# ---------------------------------------------------------------------------

def build_meeting_prep_html(company: str, research: dict) -> str:
    """Page 2: Pain points, quotes, deal context, objections, contacts, tech stack.
    All data from research JSON — no API calls.
    """
    today = date.today().strftime("%B %d, %Y")
    gong  = research.get("gong_insights") or {}
    slack_list = research.get("slack_insights") or []
    slack = slack_list[0] if slack_list else {}

    sections = []

    # ── Pain Points ──
    pain_points = gong.get("pain_points") or []
    if pain_points:
        items = "".join(f"<li>{esc(str(p))}</li>" for p in pain_points[:6])
        sections.append(_section("!", "Key Pain Points", f'<ul class="bullets">{items}</ul>'))

    # ── Verbatim Quotes ──
    quotes = gong.get("verbatim_quotes") or []
    call_date   = gong.get("call_date") or ""
    interviewee = gong.get("interviewee") or gong.get("call_title") or ""
    if quotes:
        blocks = ""
        for q in quotes[:3]:
            text = q if isinstance(q, str) else (q.get("text") or q.get("quote") or "")
            src  = q.get("source", "") if isinstance(q, dict) else f"{interviewee} — {call_date}".strip(" — ")
            if text:
                src_tag = f'<div class="quote-src">{esc(src)}</div>' if src else ""
                blocks += f'<div class="quote">&#8220;{esc(text)}&#8221;{src_tag}</div>'
        if blocks:
            sections.append(_section("❝", "Lead With", blocks))

    # ── Deal Context ──
    deal_rows = []
    for key, label in [
        ("deal_stage",        "Deal Stage"),
        ("primary_contact",   "Primary Contact"),
        ("secondary_contact", "Secondary Contact"),
        ("next_steps",        "Next Steps"),
    ]:
        if slack.get(key):
            deal_rows.append((label, str(slack[key])))
    if deal_rows:
        rows_html = "".join(
            f'<div class="deal-row"><span class="deal-label">{esc(k)}</span><span>{esc(v)}</span></div>'
            for k, v in deal_rows
        )
        sections.append(_section("◑", "Deal Context", f'<div class="deal-box">{rows_html}</div>'))

    # ── Known Objections ──
    objections = gong.get("key_objections") or []
    if objections:
        items = "".join(f"<li>{esc(str(o))}</li>" for o in objections)
        sections.append(_section("⚑", "Known Objections", f'<ul class="bullets">{items}</ul>'))

    # ── Key Contacts ──
    champions = gong.get("champions") or []
    if champions:
        rows_html = ""
        for ch in champions:
            if isinstance(ch, dict):
                rows_html += (
                    f'<div class="champion-row">'
                    f'<div class="champ-name">{esc(ch.get("name",""))}</div>'
                    f'<div class="champ-title">{esc(ch.get("title",""))}</div>'
                    f'<div class="champ-note">{esc(ch.get("note",""))}</div>'
                    f'</div>'
                )
        if rows_html:
            sections.append(_section("✦", "Key Contacts", f'<div class="contacts-box">{rows_html}</div>'))

    # ── Tech Stack ──
    tech_stack = gong.get("tech_stack") or {}
    if tech_stack and isinstance(tech_stack, dict):
        rows_html = "".join(
            f"<tr><td>{esc(k)}</td><td>{esc(v)}</td></tr>"
            for k, v in tech_stack.items()
        )
        sections.append(_section("⚙", "Tech Stack", f'<table class="tech-table">{rows_html}</table>'))

    if not sections:
        sections.append('<p style="color:#666;padding:16px 0;font-size:11px;">No meeting prep data — run /deck-research first.</p>')

    body = "\n".join(sections)

    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>{CSS}</style></head>'
        f'<body>'
        f'<div class="banner"><div>'
        f'<div class="banner-eyebrow">JOLLY.COM &middot; CONFIDENTIAL</div>'
        f'<div class="banner-title">{esc(company)}</div>'
        f'<div class="banner-sub">Meeting Prep &middot; {today}</div>'
        f'</div></div>'
        f'<div class="body">{body}'
        f'<div class="footer">'
        f'<span>Jolly.com</span><span>Confidential — Internal Use Only</span><span>{today}</span>'
        f'</div>'
        f'</div></body></html>'
    )


# ---------------------------------------------------------------------------
# Campaign Cheat Sheet
# ---------------------------------------------------------------------------

def build_campaign_html(company: str, research: dict, assumptions: dict) -> str:
    """Campaign cards sourced from research JSON campaign_details + campaigns_selected.

    assumptions: dict from read_model_assumptions() — keyed by "assumptions__{slug}".
    Pass {} if Excel model not available.
    """
    today    = date.today().strftime("%B %d, %Y")
    campaigns = research.get("campaigns_selected") or []
    details   = research.get("campaign_details") or {}

    def _get_detail(name: str) -> dict:
        if name in details:
            return details[name]
        name_lo = name.lower()
        for k, v in details.items():
            if k.lower() == name_lo:
                return v
        return {}

    def _get_assumptions(name: str) -> list:
        slug = slugify(name)
        rows = assumptions.get(f"assumptions__{slug}")
        if rows:
            return rows
        for k, v in assumptions.items():
            if k.startswith("assumptions__") and slug[:8] in k:
                return v
        return []

    normalised = []
    for c in campaigns:
        if isinstance(c, str):
            name = c
            d = _get_detail(name)
            normalised.append({
                "name": name, "priority": "standard",
                "evidence": d.get("description") or "",
                "client_interest": "",
                "rops": d.get("rops_base"),
                "ebitda": d.get("ebitda_uplift_base"),
            })
        elif isinstance(c, dict):
            name = c.get("campaign_type") or c.get("name") or ""
            d = _get_detail(name)
            normalised.append({
                "name": name,
                "priority": c.get("priority") or "standard",
                "evidence": c.get("evidence") or c.get("evidence_source") or d.get("description") or "",
                "client_interest": c.get("client_interest") or c.get("interest") or "",
                "rops": d.get("rops_base") or c.get("rops"),
                "ebitda": d.get("ebitda_uplift_base") or c.get("ebitda_impact"),
            })

    # ── Opportunity banner ──
    total_ebitda = sum(
        c["ebitda"] for c in normalised
        if isinstance(c.get("ebitda"), (int, float)) and c["ebitda"] > 0
    )
    rops_vals = [c["rops"] for c in normalised if isinstance(c.get("rops"), (int, float)) and c["rops"] > 0]
    avg_rops  = sum(rops_vals) / len(rops_vals) if rops_vals else None
    high_count = sum(1 for c in normalised if c["priority"].lower() == "high")

    opp_items = ""
    if total_ebitda:
        opp_items += f'<div class="opp-item"><div class="opp-v">{fmt(total_ebitda)}</div><div class="opp-l">Total EBITDA</div></div>'
    if avg_rops:
        opp_items += f'<div class="opp-item"><div class="opp-v">{avg_rops:.0f}x</div><div class="opp-l">Avg ROPS</div></div>'
    opp_items += f'<div class="opp-item"><div class="opp-v">{high_count}</div><div class="opp-l">High Priority</div></div>'
    opp_items += f'<div class="opp-item"><div class="opp-v">{len(normalised)}</div><div class="opp-l">Campaigns</div></div>'
    opp_banner = f'<div class="opp-banner">{opp_items}</div>' if normalised else ""

    # ── Campaign cards ──
    if not normalised:
        cards_html = '<p style="color:#666;padding:16px 0;font-size:11px;">No campaign data. Run /deck-research first.</p>'
    else:
        cards = []
        for c in normalised:
            name     = c["name"]
            priority = c["priority"].lower()
            is_high  = priority == "high"
            rops     = c["rops"]
            ebitda   = c["ebitda"]
            evidence = c["evidence"]
            interest = c["client_interest"]

            head_cls = "c-head high" if is_high else "c-head"
            interest_badge = f'<span class="badge b-interest">{esc(str(interest).capitalize())}</span>' if interest else ""
            priority_badge = f'<span class="badge {"b-high" if is_high else "b-std"}">{"High Priority" if is_high else priority.capitalize()}</span>'

            rops_fmt  = f"{rops:.0f}x" if isinstance(rops, (int, float)) else (str(rops) if rops else "—")
            ebitda_fmt = fmt(ebitda) if ebitda else "—"

            stats_html = (
                f'<div class="c-stats">'
                f'<div class="c-stat-item"><div class="c-stat-v">{esc(rops_fmt)}</div><div class="c-stat-l">ROPS</div></div>'
                f'<div class="c-stat-item"><div class="c-stat-v">{esc(ebitda_fmt)}</div><div class="c-stat-l">Est. EBITDA</div></div>'
                f'</div>'
            )
            ev_block = f'<div class="evidence">&#8220;{esc(str(evidence))}&#8221;</div>' if evidence else ""

            # Assumptions
            assump_rows = _get_assumptions(name)
            assump_block = ""
            if assump_rows:
                has_scenarios = len(assump_rows[0]) == 4 and any(r[2] or r[3] for r in assump_rows)
                if has_scenarios:
                    hdr = '<tr class="assump-hdr"><th>Assumption</th><th>Base</th><th>Upside</th><th>Downside</th></tr>'
                    row_cells = "".join(
                        f'<tr><td>{esc(lbl)}</td>'
                        f'<td class="col-base">{esc(b)}</td>'
                        f'<td class="col-up">{esc(u) if u else "—"}</td>'
                        f'<td class="col-dn">{esc(d) if d else "—"}</td></tr>'
                        for lbl, b, u, d in assump_rows
                    )
                    label_txt = "Model Assumptions — Base / Upside / Downside"
                else:
                    hdr = ""
                    row_cells = "".join(
                        f'<tr><td>{esc(lbl)}</td><td class="col-base">{esc(b)}</td></tr>'
                        for lbl, b, *_ in assump_rows
                    )
                    label_txt = "Model Assumptions (Base)"
                tbl = f'<table class="assump-table">{hdr}{row_cells}</table>'
                assump_block = f'<div class="assumptions"><div class="assumptions-label">{label_txt}</div>{tbl}</div>'

            cards.append(
                f'<div class="campaign">'
                f'<div class="{head_cls}">'
                f'<div class="c-title">{esc(name)}</div>'
                f'<div class="c-badges">{interest_badge}{priority_badge}</div>'
                f'</div>'
                f'<div class="c-body">{stats_html}{ev_block}{assump_block}</div>'
                f'</div>'
            )
        cards_html = "\n".join(cards)

    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>{CSS}</style></head>'
        f'<body>'
        f'<div class="banner"><div>'
        f'<div class="banner-eyebrow">JOLLY.COM &middot; CONFIDENTIAL</div>'
        f'<div class="banner-title">{esc(company)}</div>'
        f'<div class="banner-sub">Campaign Overview &middot; {today}</div>'
        f'</div></div>'
        f'<div class="body">'
        f'{opp_banner}{cards_html}'
        f'<div class="footer">'
        f'<span>Jolly.com</span><span>Confidential — Internal Use Only</span><span>{today}</span>'
        f'</div>'
        f'</div></body></html>'
    )


# ---------------------------------------------------------------------------
# PDF rendering
# ---------------------------------------------------------------------------

def render_with_weasyprint(html: str, pdf_path: str) -> bool:
    try:
        weasyprint.HTML(string=html).write_pdf(pdf_path)
        return True
    except Exception as e:
        print(f"  WeasyPrint error: {e}")
        return False


def render_with_playwright(html: str, pdf_path: str) -> bool:
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            try:
                page = browser.new_page(viewport={"width": 816, "height": 1056})
                page.set_content(html, wait_until="networkidle")
                page.pdf(
                    path=pdf_path,
                    format="Letter",
                    print_background=True,
                    margin={"top": "0", "bottom": "0", "left": "0", "right": "0"},
                )
            finally:
                browser.close()
        return True
    except Exception as e:
        print(f"  Playwright error: {e}")
        return False


def render_pdf(html: str, pdf_path: str) -> bool:
    """Try WeasyPrint → Playwright → HTML fallback. Never hard-fails."""
    if WEASYPRINT_AVAILABLE:
        print("  Renderer: WeasyPrint")
        if render_with_weasyprint(html, pdf_path):
            return True
    if PLAYWRIGHT_AVAILABLE:
        print("  Renderer: Playwright (fallback)")
        if render_with_playwright(html, pdf_path):
            return True
    html_path = str(pdf_path).replace(".pdf", ".html")
    Path(html_path).write_text(html, encoding="utf-8")
    print(f"  No PDF renderer available. Saved HTML: {html_path}")
    print(f"  Open in browser and Ctrl+P → Save as PDF")
    print(f"  To install: pip install weasyprint")
    return False


def _combine_pages(*page_htmls: str) -> str:
    """Combine multiple page HTML docs into a single printable document."""
    bodies = []
    for html in page_htmls:
        start = html.find("<body>")
        end   = html.rfind("</body>")
        if start != -1 and end != -1:
            bodies.append(html[start + 6:end].strip())
    css_start = page_htmls[0].find("<style>")
    css_end   = page_htmls[0].find("</style>") + 8
    css_block = page_htmls[0][css_start:css_end] if css_start != -1 else f"<style>{CSS}</style>"
    combined_body = '\n<div class="page-break"></div>\n'.join(bodies)
    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8">{css_block}</head>'
        f'<body>{combined_body}</body></html>'
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate cheat sheets for intro deck")
    parser.add_argument("--company", required=True, help="Company name (must match client folder)")
    parser.add_argument("--client-path", default=None, dest="client_path",
                        help="Override client folder path")
    args = parser.parse_args()
    company = args.company

    print(f"\n=== cheatsheet_gen.py | {company} ===\n")

    if not OPENPYXL_AVAILABLE:
        print("WARNING: openpyxl not installed — assumptions table will be empty.")
        print("  Run: pip install openpyxl\n")

    cfg = get_workspace_config()
    client_root = cfg.get("client_root", "Clients")
    client_base = Path(args.client_path) if args.client_path else Path(client_root) / company

    # ── Load research JSON ──
    try:
        research = find_research_json(company, base_path=str(client_base))
        print(f"  Research date: {research.get('research_date', 'unknown')}")
    except FileNotFoundError as e:
        print(f"  WARNING: {e}")
        print("  Generating with empty data — run /deck-research first.")
        research = {
            "company_name": company, "company_basics": {}, "attio_insights": {},
            "gong_insights": {}, "slack_insights": [], "campaigns_selected": [],
            "campaign_details": {}, "comps_benchmarks": {}, "source_summary": {},
        }

    # ── Load assumptions from Excel (optional) ──
    assumptions = {}
    if OPENPYXL_AVAILABLE:
        try:
            model_path = find_model(company, base_path=str(client_base))
            assumptions = read_model_assumptions(model_path)
            print(f"  Model: {Path(model_path).name} ({len(assumptions)} assumption groups)")
        except FileNotFoundError as e:
            print(f"  WARNING: {e} — assumptions table will be empty.")

    # ── Build HTML pages ──
    out_dir = client_base / "4. Reports" / "Cheat Sheets"
    out_dir.mkdir(parents=True, exist_ok=True)

    company_html  = build_company_html(company, research)
    meeting_html  = build_meeting_prep_html(company, research)
    campaign_html = build_campaign_html(company, research, assumptions)
    combined_html = _combine_pages(company_html, meeting_html, campaign_html)

    pdf_path = out_dir / f"{company} Cheat Sheet.pdf"
    print("\nRendering Cheat Sheet...")
    ok = render_pdf(combined_html, str(pdf_path))

    if ok:
        print(f"\n  Saved: {pdf_path}")
        print(f'\n  Open: start "" "{pdf_path}"')
    print("\nDone.")


if __name__ == "__main__":
    main()
