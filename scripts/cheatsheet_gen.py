"""
cheatsheet_gen.py — Generate company and campaign cheat sheets as a combined PDF.

Uses Playwright (headless Chrome) for pixel-perfect rendering.

Usage:
    python .claude/scripts/cheatsheet_gen.py --company "Company Name"

Output:
    Clients/[Company]/4. Reports/[Company] Cheat Sheet.pdf

Templates saved to:
    Templates/Cheat Sheets/

Requires: playwright, openpyxl
    pip install playwright openpyxl && playwright install chromium
"""
from __future__ import annotations

import sys
import json
import glob
import argparse
import os
import re
from pathlib import Path
from datetime import date

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ---------------------------------------------------------------------------
# Jolly brand (matched to jolly.com)
# ---------------------------------------------------------------------------

NAVY    = "#123769"
NAVY_LT = "#1a4a8a"
GOLD    = "#E8A838"
GOLD_LT = "#FEF6E4"
BG      = "#F6F8FA"
WHITE   = "#FFFFFF"
GRAY_LT = "#EEF1F5"
GRAY    = "#D1D5DC"
BORDER  = "#DDE2EA"
TEXT    = "#1a1a2e"
MUTED   = "#666D80"
MUTED2  = "#818898"
GREEN   = "#1a7a4a"
GREEN_LT= "#e8f6ee"
RED     = "#b03030"
RED_LT  = "#fdeaea"


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def slugify(s: str) -> str:
    return re.sub(r"[^a-z0-9_]", "", s.lower().replace(" ", "_").replace("-", "_"))



def _match_model_slug(research_slug: str, model_slugs: list[str]) -> str | None:
    """Return the best-matching model slug for a research campaign slug.

    First tries exact match, then falls back to difflib similarity.
    Returns None if no match is close enough (ratio < 0.55).
    """
    from difflib import SequenceMatcher
    if research_slug in model_slugs:
        return research_slug
    best, best_r = None, 0.0
    for ms in model_slugs:
        r = SequenceMatcher(None, research_slug, ms).ratio()
        if r > best_r:
            best_r, best = r, ms
    return best if best_r >= 0.55 else None


def esc(s) -> str:
    return (str(s or "")
            .replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def fmt(v, prefix="$") -> str:
    """Format a numeric value with optional prefix (default '$').

    Use prefix="" for plain (no-currency) formatting.
    """
    if v is None:
        return "N/A"
    try:
        v = float(v)
    except (TypeError, ValueError):
        return str(v) or "N/A"
    if 0 < abs(v) < 1:
        return f"{v * 100:.1f}%"
    if abs(v) >= 1_000_000_000:
        return f"{prefix}{v / 1_000_000_000:.1f}B"
    if abs(v) >= 1_000_000:
        return f"{prefix}{v / 1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"{prefix}{v / 1_000:.1f}K"
    if v == int(v):
        return f"{prefix}{int(v)}"
    # With prefix: fixed 2 decimals ($1.50). Without: natural repr (1.5).
    return f"{prefix}{v:.2f}" if prefix else str(round(v, 2))


def fmt_plain(v) -> str:
    return fmt(v, prefix="")


def fmt_bench(path_val) -> str:
    """Format a comps_benchmarks sub-dict {low, mid, high} as 'mid (low–high)'."""
    if isinstance(path_val, dict):
        lo  = path_val.get("low")
        mid = path_val.get("mid")
        hi  = path_val.get("high")
        if mid is not None:
            mid_s  = fmt_plain(mid)
            range_s = f"  ({fmt_plain(lo)} – {fmt_plain(hi)})" if lo is not None and hi is not None else ""
            return f"{mid_s}{range_s}"
        return str(path_val)
    return fmt_plain(path_val)


def _deep_get(d: dict, path: str, default=None):
    """Traverse nested dict/list via dot-separated path (supports integer indices)."""
    parts = path.split(".")
    cur = d
    for p in parts:
        if isinstance(cur, dict):
            cur = cur.get(p)
        elif isinstance(cur, list) and p.isdigit():
            idx = int(p)
            cur = cur[idx] if idx < len(cur) else None
        else:
            return default
        if cur is None:
            return default
    return cur


def get_workspace_config() -> dict:
    candidates = [".claude/data/workspace_config.json"]
    jolly_ws = os.environ.get("JOLLY_WORKSPACE", "").strip().strip("\r")
    if jolly_ws:
        candidates.append(os.path.join(jolly_ws, ".claude/data/workspace_config.json"))
    for p in candidates:
        if os.path.exists(p):
            try:
                with open(p, encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    return {}


def find_research_json(company: str, base_path: str = None) -> dict:
    slug = slugify(company)

    if base_path:
        client_glob = f"{base_path}/**/research_output_*.json"
    else:
        cfg = get_workspace_config()
        client_root = cfg.get("client_root", "Clients")
        client_glob = f"{client_root}/{company}/**/research_output_*.json"

    candidates = []
    for m in glob.glob(client_glob, recursive=True):
        try:
            candidates.append((os.path.getmtime(m), m))
        except OSError:
            pass
    for _, path in sorted(candidates, reverse=True):
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            print(f"  Research JSON: {path}")
            return data
        except (json.JSONDecodeError, OSError):
            continue
    for p in [
        f".claude/data/research_output_{slug}.json",
        f".claude/data/research_output_{company.lower().replace(' ', '-')}.json",
    ]:
        if os.path.exists(p):
            print(f"  Research JSON (legacy): {p}")
            with open(p, encoding="utf-8") as f:
                return json.load(f)
    raise FileNotFoundError(f"No research JSON found for '{company}'")


def find_model(company: str, base_path: str = None) -> str:
    if base_path:
        model_glob = f"{base_path}/1. Model/*.xlsx"
        search_path = f"{base_path}/1. Model/"
    else:
        cfg = get_workspace_config()
        client_root = cfg.get("client_root", "Clients")
        model_glob = f"{client_root}/{company}/1. Model/*.xlsx"
        search_path = f"{client_root}/{company}/1. Model/"

    matches = glob.glob(model_glob)
    if not matches:
        raise FileNotFoundError(f"No Excel model in {search_path}")
    return sorted(matches)[-1]


def _call_claude_api(prompt: str, max_tokens: int = 1024, timeout: int = 30) -> dict | None:
    """Call Claude API and return parsed JSON, or None on failure.

    Handles: API key check, HTTP request, markdown fence stripping, JSON parse.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return None
    import urllib.request
    payload = json.dumps({
        "model": "claude-haiku-4-5-20251001",
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=payload,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        body = json.loads(resp.read().decode("utf-8"))
    text = body["content"][0]["text"].strip()
    # Strip any accidental markdown fences
    if text.startswith("```"):
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
    return json.loads(text)


def _generate_config_with_claude(research: dict, vertical: str) -> dict | None:
    """Call Claude API to generate a report config for an unrecognised vertical."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return None

    basics = research.get("company_basics") or {}
    comps  = research.get("comps_benchmarks") or {}

    # Summarise what data is actually available so Claude can make sensible choices
    available_basics = [k for k, v in basics.items() if v not in (None, {}, [], "")]
    available_comps  = [k for k, v in comps.items()
                        if isinstance(v, (int, float, dict)) and k not in ("vertical", "stale", "cache_last_refreshed")]
    geo_keys   = list((basics.get("geography") or {}).keys())
    has_breakdown = bool(basics.get("employee_breakdown"))
    breakdown_keys = list((basics.get("employee_breakdown") or {}).keys())

    prompt = f"""You are generating a JSON report config for a Jolly sales cheat sheet tool.

Industry: {research.get("industry", vertical)}
Vertical slug: {vertical}

Available company_basics fields: {available_basics}
Available comps_benchmarks fields: {available_comps}
Geography sub-fields: {geo_keys}
Has employee_breakdown: {has_breakdown}
Breakdown keys: {breakdown_keys}

Generate a JSON object with exactly these keys:
- "vertical": "{vertical}"
- "unit_label": string — what to call unit_count for this industry (e.g. "Restaurants", "Facilities", "Service Centers", "Stores", "Properties", "Clinics")
- "profile_fields": array of objects, each with "label" (string), "path" (dot-notation into research JSON), and optionally "format" ("text"|"count"|"currency"|"percent"). Include industry-relevant fields. Use paths like "industry", "attio_insights.founded", "company_basics.geography.hq", "company_basics.geography.rank", "company_basics.geography.state_count", and any of the available_basics fields prefixed with "company_basics.". Only include fields that make sense for this industry.
- "employee_breakdown": {str(has_breakdown).lower()} (boolean)
- "employee_breakdown_keys": array of breakdown key strings (from breakdown_keys above, if employee_breakdown is true)
- "employee_breakdown_labels": array of human-readable labels matching each key
- "benchmark_fields": array of objects with "label" and "path". Use paths like "comps_benchmarks.ebitda_margin_pct", "comps_benchmarks.turnover_rate", "comps_benchmarks.gross_margin_pct" — only include ones from available_comps.

Respond with only the JSON object, no markdown fences, no explanation."""

    try:
        config = _call_claude_api(prompt, max_tokens=1024, timeout=30)
        if config is None:
            return None
        config["_auto_generated"] = True
        config["_note"] = "Auto-generated by Claude. Edit to customise which fields appear."
        return config
    except Exception as e:
        print(f"  WARNING: Claude config generation failed: {e}")
        return None


def _generate_config_fallback(research: dict, vertical: str) -> dict:
    """Programmatic fallback: build a config from what's present in the research JSON."""
    basics   = research.get("company_basics") or {}
    comps    = research.get("comps_benchmarks") or {}
    industry = (research.get("industry") or vertical).lower()

    # Unit label heuristic
    if any(k in industry for k in ["restaurant", "qsr", "food service", "fast food"]):
        unit_label = "Restaurants"
    elif any(k in industry for k in ["manufactur", "plant", "production", "bottl"]):
        unit_label = "Facilities"
    elif any(k in industry for k in ["automotive", "auto service", "dealership"]):
        unit_label = "Service Centers"
    elif any(k in industry for k in ["retail", "store", "shop"]):
        unit_label = "Stores"
    elif any(k in industry for k in ["hotel", "hospitality", "lodging", "resort"]):
        unit_label = "Properties"
    elif any(k in industry for k in ["clinic", "healthcare", "medical", "dental"]):
        unit_label = "Clinics"
    else:
        unit_label = "Locations"

    # Build profile fields — always include the core set, then any populated extras
    attio = research.get("attio_insights") or {}
    slack_list = research.get("slack_insights") or []
    slack = slack_list[0] if slack_list else {}

    profile_fields = [
        {"label": "Industry",        "path": "industry"},
        {"label": "Founded",         "path": "attio_insights.founded"},
        {"label": "HQ",              "path": "company_basics.geography.hq"},
        {"label": "Market Position", "path": "company_basics.geography.rank"},
        {"label": "States Operated", "path": "company_basics.geography.state_count"},
    ]

    # Attio details
    if attio.get("categories"):
        profile_fields.append({"label": "Categories", "path": "attio_insights.categories"})
    if attio.get("strongest_connection_strength"):
        profile_fields.append({"label": "Connection", "path": "attio_insights.strongest_connection_strength"})
    if attio.get("employee_range"):
        profile_fields.append({"label": "Employee Range", "path": "attio_insights.employee_range"})
    if attio.get("estimated_arr_usd"):
        profile_fields.append({"label": "Est. ARR", "path": "attio_insights.estimated_arr_usd"})
    if attio.get("last_interaction"):
        profile_fields.append({"label": "Last Interaction", "path": "attio_insights.last_interaction"})

    # Slack insights
    if slack.get("deal_stage"):
        profile_fields.append({"label": "Deal Stage", "path": "slack_insights.0.deal_stage"})
    if slack.get("primary_contact"):
        profile_fields.append({"label": "Primary Contact", "path": "slack_insights.0.primary_contact"})
    if slack.get("secondary_contact"):
        profile_fields.append({"label": "Secondary Contact", "path": "slack_insights.0.secondary_contact"})
    if slack.get("next_steps"):
        profile_fields.append({"label": "Next Steps", "path": "slack_insights.0.next_steps"})
    EXTRA_FIELD_MAP = {
        "annual_case_volume": ("Annual Case Volume", "count"),
        "customer_count":     ("Customer Count",     "count"),
        "avg_unit_volume":    ("Avg Unit Volume",     "currency"),
        "technician_count":   ("Technician Count",    "count"),
        "annual_job_count":   ("Jobs / Year",         "count"),
        "daypart_mix":        ("Daypart Mix",          "text"),
        "recent_capex":       ("Recent CapEx",         "text"),
    }
    for key, (label, fmt_mode) in EXTRA_FIELD_MAP.items():
        if basics.get(key) is not None:
            profile_fields.append({"label": label, "path": f"company_basics.{key}", "format": fmt_mode})

    has_breakdown  = bool(basics.get("employee_breakdown"))
    breakdown_keys = list((basics.get("employee_breakdown") or {}).keys())
    breakdown_lbls = [k.replace("_", " ").title() for k in breakdown_keys]

    BENCH_MAP = [
        ("ebitda_margin_pct", "EBITDA Margin"),
        ("gross_margin_pct",  "Gross Margin"),
        ("turnover_rate",     "Turnover Rate"),
        ("net_margin_pct",    "Net Margin"),
        ("labor_cost_pct",    "Labor Cost %"),
    ]
    benchmark_fields = [
        {"label": lbl, "path": f"comps_benchmarks.{key}"}
        for key, lbl in BENCH_MAP
        if comps.get(key) is not None
    ] or [
        {"label": "EBITDA Margin", "path": "comps_benchmarks.ebitda_margin_pct"},
        {"label": "Turnover Rate", "path": "comps_benchmarks.turnover_rate"},
    ]

    return {
        "vertical":                vertical,
        "unit_label":              unit_label,
        "profile_fields":          profile_fields,
        "employee_breakdown":      has_breakdown,
        "employee_breakdown_keys": breakdown_keys,
        "employee_breakdown_labels": breakdown_lbls,
        "benchmark_fields":        benchmark_fields,
        "_auto_generated":         True,
        "_note": "Auto-generated (fallback). Edit to customise which fields appear.",
    }


def load_vertical_config(research: dict) -> dict:
    """Load vertical report config, auto-generating and saving one if none exists."""
    cfg = get_workspace_config()
    templates_root = cfg.get("templates_root", "Templates")
    tpl_dir = Path(templates_root) / "Cheat Sheets"

    # Detect vertical
    vertical = (
        (research.get("comps_benchmarks") or {}).get("vertical") or
        slugify((research.get("industry") or "").split("/")[0])
    ) or "default"

    # 1. Try exact match for this vertical
    exact = tpl_dir / f"report_config_{vertical}.json"
    if exact.exists():
        try:
            with open(exact, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    # 2. No config found — auto-generate with Claude, fall back to programmatic
    print(f"  No report config for vertical '{vertical}' — generating with Claude...")
    config = _generate_config_with_claude(research, vertical)
    if config is None:
        print("  Falling back to programmatic config generation.")
        config = _generate_config_fallback(research, vertical)

    # 3. Save so it's reusable and editable
    tpl_dir.mkdir(parents=True, exist_ok=True)
    try:
        with open(exact, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
        source = "Claude" if config.get("_auto_generated") and "fallback" not in config.get("_note","") else "auto-generated"
        print(f"  New config saved ({source}): {exact}")
        print(f"  Edit it to customise which fields appear on the Company Cheat Sheet.")
    except Exception as e:
        print(f"  WARNING: Could not save config: {e}")

    # 4. Last resort: load default if something went wrong
    default = tpl_dir / "report_config_default.json"
    if not config and default.exists():
        with open(default, encoding="utf-8") as f:
            return json.load(f)

    return config


def _compute_total_opportunity(model_values: dict, normalised_campaigns: list) -> dict:
    """Compute total EBITDA, average ROPS, and campaign counts from model values."""
    ebitda_vals = [(k, v) for k, v in model_values.items()
                   if k.startswith("ebitda__") and isinstance(v, (int, float)) and v > 0]
    rops_vals   = [v for k, v in model_values.items()
                   if k.startswith("rops__") and isinstance(v, (int, float)) and v > 0]
    total_ebitda = sum(v for _, v in ebitda_vals)
    avg_rops     = sum(rops_vals) / len(rops_vals) if rops_vals else None
    high_count   = sum(1 for c in normalised_campaigns if c.get("priority","").lower() == "high")
    return {
        "total_ebitda": total_ebitda,
        "avg_rops":     avg_rops,
        "high_count":   high_count,
        "total_count":  len(normalised_campaigns),
    }


def _generate_meeting_intelligence(company: str, research: dict, model_values: dict,
                                    cache_dir: str = "") -> dict:
    """Call Claude to generate meeting-prep insights. Caches result for 7 days."""
    # Load from cache if fresh
    if cache_dir:
        cache_path = Path(cache_dir) / f"meeting_intel_{slugify(company)}.json"
        if cache_path.exists():
            try:
                from datetime import date as _date
                age = (_date.today().toordinal() -
                       _date.fromtimestamp(cache_path.stat().st_mtime).toordinal())
                if age < 7:
                    with open(cache_path, encoding="utf-8") as f:
                        data = json.load(f)
                    print(f"  Meeting intelligence: from cache ({age}d old)")
                    return data
            except Exception:
                pass

    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        return {}

    basics  = research.get("company_basics") or {}
    gong    = research.get("gong_insights") or {}
    camps   = research.get("campaigns_selected") or []
    geo     = basics.get("geography") or {}

    opp = _compute_total_opportunity(model_values, [])
    ebitda_total = opp["total_ebitda"]
    avg_rops     = opp["avg_rops"]

    camp_names  = [c.get("name","") if isinstance(c,dict) else str(c) for c in camps]
    high_camps  = [c.get("name","") if isinstance(c,dict) else str(c) for c in camps
                   if isinstance(c,dict) and (c.get("priority","") or "").lower() == "high"]
    tech_lines  = "\n".join(f"- {k}: {v}" for k,v in (gong.get("tech_stack") or {}).items())
    obj_lines   = "\n".join(f"- {o}" for o in (gong.get("key_objections") or []))
    pain_lines  = "\n".join(f"- {p}" for p in (gong.get("pain_points") or []))
    quote_lines = "\n".join(f'- "{q}"' for q in (gong.get("verbatim_quotes") or [])[:4])
    opp_str = f"${ebitda_total/1e6:.1f}M EBITDA across {len(camp_names)} campaigns" + \
              (f", avg ROPS {avg_rops:.0f}x" if avg_rops else "")

    prompt = f"""You are a senior sales strategist at Jolly, preparing a rep for a first meeting with {company}.

Company: {company} | Industry: {research.get("industry","")}
Revenue: ${basics.get("annual_revenue",0)/1e9:.1f}B | Employees: {basics.get("employee_count","?")}
Geography: {geo.get("rank","")}, {geo.get("state_count","?")} states

Pain points from Gong:
{pain_lines}

Verbatim quotes from executive call:
{quote_lines}

Current tools:
{tech_lines}

Key objections:
{obj_lines}

Campaigns selected: {", ".join(camp_names)}
High priority: {", ".join(high_camps)}
Total modeled opportunity: {opp_str}

Generate a JSON object with these exact keys:
- "headline_insights": array of exactly 3 strings — the most important things to know walking in. Be specific to this company; cite actual data, quotes, or situations. NOT generic.
- "competitive_angle": string — 1-2 sentences on what makes Jolly better than what they currently use. Reference their actual tools. Be concrete and direct.
- "conversation_starters": array of exactly 3 strings — specific openers that show you've done your homework on this company. Reference real facts about them.
- "deal_dynamics": array of 3-4 strings — what to expect in this specific sales process. Reference their stated requirements (pilot, ROI window, etc.).
- "objection_handling": object — key = exact objection string from the list above, value = 1-sentence response. Include all objections listed.

Return only valid JSON. No markdown fences."""

    try:
        result = _call_claude_api(prompt, max_tokens=2000, timeout=45)
        if result is None:
            return {}
        # Cache
        if cache_dir:
            try:
                Path(cache_dir).mkdir(parents=True, exist_ok=True)
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(result, f, indent=2)
                print(f"  Meeting intelligence: generated by Claude (cached)")
            except Exception:
                pass
        return result
    except Exception as e:
        print(f"  WARNING: Meeting intelligence generation failed: {e}")
        return {}


def _fmt_assumption(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if 0 < val < 1:
            return f"{val:.0%}"
        if val >= 1000:
            return f"${val:,.0f}"
        if val >= 100:
            return f"{val:,.0f}"
        if val == int(val):
            return f"{int(val):,}"
        return f"{val:.2f}"
    if isinstance(val, int):
        if val >= 1000:
            return f"${val:,}"
        return f"{val:,}"
    return str(val)


def read_model_basics(model_path: str) -> dict:
    if not OPENPYXL_AVAILABLE:
        return {}
    # If the file is locked by Excel, copy it to a temp location first
    try:
        wb = load_workbook(model_path, data_only=True)
    except PermissionError:
        import shutil, tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        try:
            shutil.copy2(model_path, tmp_path)
            wb = load_workbook(tmp_path, data_only=True)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
    values = {}
    campaign_slots = {}

    if "Inputs" in wb.sheetnames:
        ws = wb["Inputs"]
        all_rows = list(ws.iter_rows(min_row=1, max_row=200, values_only=True))

        for row in all_rows:
            label = row[1] if len(row) > 1 else None
            val   = row[4] if len(row) > 4 else None
            if label and val is not None:
                values[str(label).strip()] = val
            b, c = (row[1] if len(row) > 1 else None), (row[2] if len(row) > 2 else None)
            if b and str(b).strip().startswith("Campaign ") and c:
                slot = str(b).strip()
                if not any(sub in slot for sub in ["3a", "3b", "3c"]):
                    campaign_slots[slot] = str(c).strip()

        scenario_start = None
        for i, row in enumerate(all_rows):
            if row[1] and "SCENARIO ASSUMPTIONS" in str(row[1]):
                scenario_start = i + 1
                break

        if scenario_start is not None:
            # Parse assumption groups keyed by their campaign header name.
            # Each group starts with a "Campaign N: Name" header row (base=None)
            # followed by assumption rows (base != None), ended by a blank row.
            current_name = None
            current_rows = []
            for row in all_rows[scenario_start:]:
                label    = row[1] if len(row) > 1 else None
                base_val = row[2] if len(row) > 2 else None   # col C = Base
                up_val   = row[3] if len(row) > 3 else None   # col D = Upside
                dn_val   = row[4] if len(row) > 4 else None   # col E = Downside
                label_str = str(label).strip() if label else ""

                if not label_str:
                    # Blank row — save current group if any
                    if current_name and current_rows:
                        key = f"assumptions__{slugify(current_name)}"
                        values[key] = current_rows
                        current_name = None
                        current_rows = []
                    continue

                # Check if this is a campaign header (e.g. "Campaign 3: Employee Timeliness & Attendance")
                if re.match(r"Campaign \d+:", label_str) and base_val is None:
                    # Save previous group if any
                    if current_name and current_rows:
                        key = f"assumptions__{slugify(current_name)}"
                        values[key] = current_rows
                    # Extract campaign name after the colon
                    current_name = label_str.split(":", 1)[1].strip()
                    current_rows = []
                elif base_val is not None:
                    current_rows.append((
                        label_str,
                        _fmt_assumption(base_val),
                        _fmt_assumption(up_val),
                        _fmt_assumption(dn_val),
                    ))

            # Save last group
            if current_name and current_rows:
                key = f"assumptions__{slugify(current_name)}"
                values[key] = current_rows

    if "Campaigns" in wb.sheetnames:
        ws_c = wb["Campaigns"]
        rops_vals   = []
        ebitda_vals = []
        for row in ws_c.iter_rows(min_row=1, max_row=200, values_only=True):
            label = row[1] if len(row) > 1 else None
            base  = row[2] if len(row) > 2 else None
            if not label:
                continue
            label_str = str(label).strip()
            if "Return on Points Spend" in label_str and base is not None:
                rops_vals.append(base)
            if "EBITDA Uplift" in label_str and base is not None:
                ebitda_vals.append(base)

        slot_order = [v for _, v in sorted(campaign_slots.items())]
        for i, name in enumerate(slot_order):
            slug = slugify(name)
            if i < len(rops_vals):
                values[f"rops__{slug}"] = rops_vals[i]
            if i < len(ebitda_vals):
                values[f"ebitda__{slug}"] = ebitda_vals[i]

    return values


# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

CSS = f"""
@page {{ size: Letter; }}
html {{ width: 816px; overflow: hidden; }}
body {{ width: 816px; overflow: hidden; }}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}

body {{
    font-family: -apple-system, 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    font-size: 10px;
    color: {TEXT};
    background: {BG};
    line-height: 1.5;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
}}

/* ── Cover ── */
.cover {{
    background: linear-gradient(135deg, {NAVY} 0%, {NAVY_LT} 100%);
    padding: 20px 28px 17px;
    position: relative;
    overflow: hidden;
}}
.cover::before {{
    content: "";
    position: absolute;
    bottom: -50px; right: -30px;
    width: 180px; height: 180px;
    border-radius: 50%;
    background: rgba(232,168,56,0.10);
}}
.cover-eyebrow {{
    font-size: 7.5px;
    font-weight: 700;
    letter-spacing: 1.8px;
    text-transform: uppercase;
    color: {GOLD};
    margin-bottom: 5px;
}}
.cover h1 {{
    font-size: 22px;
    font-weight: 800;
    color: {WHITE};
    letter-spacing: -0.3px;
    line-height: 1.1;
    margin-bottom: 4px;
}}
.cover-sub {{
    font-size: 8.5px;
    color: rgba(255,255,255,0.5);
    letter-spacing: 0.3px;
}}
.gold-rule {{
    height: 3px;
    background: linear-gradient(90deg, {GOLD} 0%, rgba(232,168,56,0.20) 100%);
}}

/* ── Body ── */
.body {{
    padding: 8px 28px 20px;
    background: {BG};
}}

/* ── Section label ── */
.section {{
    margin: 10px 0 5px;
    break-inside: avoid;
}}
.section-label {{
    font-size: 7.5px;
    font-weight: 800;
    letter-spacing: 1.4px;
    text-transform: uppercase;
    color: {NAVY};
    padding-bottom: 4px;
    border-bottom: 2px solid {GOLD};
    display: flex;
    align-items: center;
    gap: 5px;
}}
.section-label + * {{ break-before: avoid; }}
.s-icon {{
    width: 14px; height: 14px;
    background: {GOLD};
    border-radius: 3px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    font-size: 8px;
    color: {NAVY};
    font-weight: 900;
    flex-shrink: 0;
}}

/* ── Total opportunity banner ── */
.opp-banner {{
    display: flex;
    gap: 0;
    margin: 6px 0 12px;
    border: 1px solid {GOLD};
    border-radius: 7px;
    overflow: hidden;
    background: {GOLD_LT};
}}
.opp-item {{
    flex: 1;
    padding: 10px 12px 9px;
    text-align: center;
    border-right: 1px solid rgba(232,168,56,0.35);
}}
.opp-item:last-child {{ border-right: none; }}
.opp-v {{
    font-size: 20px;
    font-weight: 900;
    color: {NAVY};
    letter-spacing: -0.5px;
    line-height: 1;
}}
.opp-l {{
    font-size: 7px;
    font-weight: 700;
    color: {MUTED};
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-top: 3px;
}}

/* ── Meeting intelligence ── */
.intel-subsection {{
    margin-top: 8px;
}}
.intel-sub-label {{
    font-size: 7px;
    font-weight: 800;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    color: {GOLD};
    margin-bottom: 4px;
    padding-bottom: 2px;
    border-bottom: 1px solid rgba(232,168,56,0.3);
}}
.intel-hl {{
    background: {GOLD_LT};
    border-left: 3px solid {GOLD};
    border-radius: 0 5px 5px 0;
    padding: 7px 11px;
    font-size: 9.5px;
    line-height: 1.55;
    margin-bottom: 5px;
}}
.intel-hl li {{
    padding: 2px 0;
}}
.intel-angle {{
    background: linear-gradient(135deg, {NAVY} 0%, {NAVY_LT} 100%);
    border-radius: 6px;
    padding: 8px 12px;
    font-size: 9.5px;
    color: {WHITE};
    line-height: 1.5;
}}
.numbered-list {{
    list-style: decimal;
    margin-top: 5px;
    padding-left: 18px;
}}
.numbered-list li {{
    padding: 3px 0;
    font-size: 9.5px;
    border-bottom: 1px solid {GRAY_LT};
}}
.numbered-list li:last-child {{ border-bottom: none; }}

/* ── KPI strip ── */
.kpi-strip {{
    display: flex;
    gap: 7px;
    margin-top: 7px;
}}
.kpi-card {{
    flex: 1;
    background: {WHITE};
    border: 1px solid {GRAY_LT};
    border-radius: 7px;
    padding: 10px 10px 8px;
    text-align: center;
    border-top: 3px solid {GOLD};
    box-shadow: 0 1px 3px rgba(18,55,105,0.06);
}}
.kpi-v {{
    font-size: 18px;
    font-weight: 800;
    color: {NAVY};
    letter-spacing: -0.5px;
    line-height: 1;
    margin-bottom: 3px;
}}
.kpi-l {{
    font-size: 7px;
    font-weight: 700;
    color: {MUTED};
    text-transform: uppercase;
    letter-spacing: 0.5px;
}}
.kpi-src {{
    font-size: 6.5px;
    color: {MUTED2};
    margin-top: 2px;
    font-style: italic;
    overflow-wrap: break-word;
    word-break: break-word;
}}

/* ── Shared table base (data, bench, tech, sum, objection) ── */
.data-table,
.bench-table,
.tech-table,
.sum-table,
.objection-table {{
    width: 100%;
    border-collapse: collapse;
    margin-top: 6px;
    background: {WHITE};
    border-radius: 6px;
    overflow: hidden;
    box-shadow: 0 1px 3px rgba(18,55,105,0.05);
}}
.data-table td,
.bench-table td,
.tech-table td,
.sum-table td,
.objection-table td {{
    padding: 5px 9px;
    border-bottom: 1px solid {GRAY_LT};
    font-size: 9px;
}}
.data-table tr:last-child td,
.bench-table tr:last-child td,
.tech-table tr:last-child td,
.sum-table tr:last-child td,
.objection-table tr:last-child td {{ border-bottom: none; }}
.data-table tr:nth-child(even) td,
.bench-table tr:nth-child(even) td,
.tech-table tr:nth-child(even) td,
.sum-table tr:nth-child(even) td,
.objection-table tr:nth-child(even) td {{ background: {BG}; }}

/* ── Objection table overrides ── */
.objection-table {{
    margin-top: 5px;
}}
.objection-table td {{
    vertical-align: top;
}}
.objection-table td:first-child {{
    font-style: italic;
    color: {MUTED};
    width: 42%;
    font-size: 8.5px;
}}
.objection-table td:last-child {{
    color: {TEXT};
}}
.objection-table tr {{ break-inside: avoid; }}

/* ── Data table ── */
.data-table th {{
    background: {NAVY};
    color: {WHITE};
    padding: 5px 9px;
    text-align: left;
    font-size: 7.5px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.4px;
}}
.data-table td {{
    vertical-align: top;
}}
.data-table td:first-child {{
    font-weight: 600;
    color: {MUTED};
    font-size: 8px;
    width: 42%;
}}

/* ── Breakdown grid ── */
.breakdown-grid {{
    display: flex;
    gap: 6px;
    margin-top: 6px;
}}
.bk-cell {{
    flex: 1;
    background: {WHITE};
    border: 1px solid {GRAY_LT};
    border-radius: 6px;
    padding: 8px 8px 7px;
    text-align: center;
    box-shadow: 0 1px 2px rgba(18,55,105,0.04);
}}
.bk-v {{
    font-size: 15px;
    font-weight: 800;
    color: {NAVY};
    letter-spacing: -0.3px;
}}
.bk-l {{
    font-size: 7px;
    font-weight: 600;
    color: {MUTED};
    text-transform: uppercase;
    letter-spacing: 0.3px;
    margin-top: 2px;
}}

/* ── Benchmark table (low/mid/high) ── */
.bench-table th {{
    background: {NAVY};
    color: {WHITE};
    padding: 5px 9px;
    font-size: 7.5px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.3px;
    text-align: center;
}}
.bench-table th:first-child {{ text-align: left; }}
.bench-table td {{
    text-align: center;
}}
.bench-table td:first-child {{
    text-align: left;
    font-weight: 600;
    color: {MUTED};
    font-size: 8px;
    width: 45%;
}}
.bench-table .mid {{ font-weight: 800; color: {NAVY}; }}

/* ── Bullet list ── */
.bullets {{ list-style: none; margin-top: 6px; }}
.bullets li {{
    padding: 4px 0 4px 14px;
    border-bottom: 1px solid {GRAY_LT};
    font-size: 9.5px;
    position: relative;
}}
.bullets li:last-child {{ border-bottom: none; }}
.bullets li::before {{
    content: "";
    position: absolute;
    left: 0; top: 10px;
    width: 5px; height: 5px;
    border-radius: 50%;
    background: {GOLD};
}}

/* ── Quote ── */
.quote {{
    background: {GOLD_LT};
    border-left: 3px solid {GOLD};
    border-radius: 0 5px 5px 0;
    padding: 7px 11px;
    margin: 4px 0;
    font-size: 9.5px;
    font-style: italic;
    line-height: 1.5;
    break-inside: avoid;
}}
.quote-src {{
    font-style: normal;
    font-size: 7.5px;
    color: {MUTED};
    margin-top: 2px;
    font-weight: 600;
}}

/* ── Champions ── */
.champion-row {{
    display: flex;
    align-items: flex-start;
    padding: 6px 0;
    border-bottom: 1px solid {GRAY_LT};
    gap: 10px;
}}
.champion-row:last-child {{ border-bottom: none; }}
.champ-name {{
    font-weight: 700;
    font-size: 9.5px;
    min-width: 160px;
}}
.champ-title {{
    font-size: 8.5px;
    color: {MUTED};
    min-width: 140px;
}}
.champ-note {{
    font-size: 8.5px;
    color: {TEXT};
    flex: 1;
}}

/* ── Tech stack ── */
.tech-table {{
    break-inside: avoid;
}}
.tech-table td {{
    vertical-align: top;
}}
.tech-table td:first-child {{
    font-weight: 700;
    color: {NAVY};
    font-size: 8.5px;
    width: 30%;
    white-space: nowrap;
}}

/* ── Source pills ── */
.pills {{ display: flex; flex-wrap: wrap; gap: 4px; margin-top: 6px; }}
.pill {{
    background: {NAVY};
    color: {WHITE};
    font-size: 7px;
    font-weight: 700;
    padding: 2px 8px;
    border-radius: 10px;
    letter-spacing: 0.3px;
}}
.pill-zero {{ background: {GRAY_LT}; color: {MUTED}; border: 1px solid {GRAY}; }}
.pill-val {{
    background: {GOLD}; color: {NAVY};
    margin-left: 2px; padding: 2px 6px;
    border-radius: 10px; font-size: 7px; font-weight: 800;
}}

/* ── Campaign cards ── */
.campaign {{
    border: 1px solid {GRAY};
    border-radius: 7px;
    margin-bottom: 8px;
    overflow: hidden;
    break-inside: avoid;
    box-shadow: 0 1px 4px rgba(18,55,105,0.07);
}}
.c-head {{
    background: linear-gradient(135deg, {NAVY} 0%, {NAVY_LT} 100%);
    padding: 9px 14px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 8px;
}}
.c-head.high {{ border-top: 3px solid {GOLD}; }}
.c-title {{
    font-size: 12px;
    font-weight: 800;
    color: {WHITE};
    letter-spacing: -0.2px;
    flex: 1;
}}
.c-badges {{
    display: flex;
    align-items: center;
    gap: 5px;
    flex-shrink: 0;
}}
.badge {{
    font-size: 7px;
    font-weight: 800;
    padding: 2px 8px;
    border-radius: 10px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    white-space: nowrap;
}}
.b-high {{ background: {GOLD}; color: {NAVY}; }}
.b-std  {{
    background: rgba(255,255,255,0.12);
    color: rgba(255,255,255,0.75);
    border: 1px solid rgba(255,255,255,0.22);
}}
.b-interest {{
    background: rgba(255,255,255,0.15);
    color: rgba(255,255,255,0.85);
    border: 1px solid rgba(255,255,255,0.30);
    font-size: 6.5px;
    font-weight: 700;
}}

/* ── Campaign body ── */
.c-body {{ padding: 9px 14px 11px; background: {WHITE}; }}

/* ── Quick stats row (ROPS / Savings / Summary) ── */
.c-stats {{
    display: flex;
    gap: 0;
    margin-bottom: 7px;
    border: 1px solid {GRAY_LT};
    border-radius: 6px;
    overflow: hidden;
    background: {BG};
}}
.c-stat-item {{
    flex: 1;
    padding: 6px 10px;
    text-align: center;
    border-right: 1px solid {GRAY_LT};
}}
.c-stat-item:last-child {{ border-right: none; }}
.c-stat-v {{
    font-size: 14px;
    font-weight: 900;
    color: {NAVY};
    letter-spacing: -0.3px;
    line-height: 1;
}}
.c-stat-l {{
    font-size: 6.5px;
    font-weight: 700;
    color: {MUTED};
    text-transform: uppercase;
    letter-spacing: 0.4px;
    margin-top: 2px;
}}

/* ── Evidence block ── */
.evidence {{
    background: {GOLD_LT};
    border-left: 3px solid {GOLD};
    border-radius: 0 5px 5px 0;
    padding: 6px 10px;
    font-size: 9px;
    font-style: italic;
    color: {TEXT};
    margin-bottom: 7px;
    line-height: 1.5;
}}

/* ── Assumptions table (supports 2-col base-only and 4-col scenarios) ── */
.assumptions {{
    margin-top: 0;
}}
.assumptions-label {{
    font-size: 7px;
    font-weight: 800;
    letter-spacing: 1.1px;
    text-transform: uppercase;
    color: {MUTED};
    margin-bottom: 4px;
}}
.assump-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 8.5px;
    background: {WHITE};
    border: 1px solid {GRAY_LT};
    border-radius: 5px;
    overflow: hidden;
}}
.assump-hdr th {{
    background: {NAVY};
    color: {WHITE};
    padding: 4px 8px;
    font-size: 7px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.3px;
    text-align: center;
}}
.assump-hdr th:first-child {{ text-align: left; }}
.assump-table td {{
    padding: 4px 8px;
    vertical-align: top;
    border-bottom: 1px solid {GRAY_LT};
}}
.assump-table tr:last-child td {{ border-bottom: none; }}
.assump-table tr:nth-child(even) td {{ background: {BG}; }}
.assump-table tr {{ break-inside: avoid; }}
.assump-table td:first-child {{
    color: {MUTED};
    font-weight: 500;
    width: 55%;
}}
.assump-table td:not(:first-child) {{
    font-weight: 700;
    text-align: right;
    white-space: nowrap;
    width: 15%;
}}
.assump-table .col-base  {{ color: {NAVY}; }}
.assump-table .col-up    {{ color: {GREEN}; background: {GREEN_LT}; }}
.assump-table .col-dn    {{ color: {RED};   background: {RED_LT}; }}

/* ── Summary table ── */
.sum-table th {{
    background: {NAVY}; color: {WHITE};
    padding: 6px 9px; text-align: left;
    font-size: 7.5px; font-weight: 700;
    text-transform: uppercase; letter-spacing: 0.4px;
}}
.sum-table td {{
    vertical-align: middle;
}}
.ch {{ background: {NAVY}; color: {WHITE}; font-size: 6.5px; font-weight: 700; padding: 2px 6px; border-radius: 8px; }}
.cs {{ background: {GRAY_LT}; color: {MUTED}; font-size: 6.5px; font-weight: 700; padding: 2px 6px; border-radius: 8px; }}

/* ── Two-col layout ── */
.two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; }}
.two-col .section {{ margin-top: 0; }}

/* ── Footer ── */
.footer {{
    margin-top: 18px;
    padding-top: 8px;
    border-top: 1px solid {GRAY_LT};
    display: flex;
    justify-content: space-between;
    font-size: 7px;
    color: {MUTED2};
}}

/* ── Page break ── */
.page-break {{ break-before: page; }}

.banner {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    height: 50px;
    background: linear-gradient(135deg, {NAVY} 0%, {NAVY_LT} 100%);
    border-bottom: 3px solid {GOLD};
    padding: 7px 28px 6px;
    box-sizing: border-box;
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
    font-family: -apple-system, 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
}}
.banner-eyebrow {{
    font-size: 6.5px;
    font-weight: 700;
    letter-spacing: 1.8px;
    text-transform: uppercase;
    color: {GOLD};
    margin-bottom: 2px;
}}
.banner-title {{
    font-size: 14px;
    font-weight: 800;
    color: #fff;
    letter-spacing: -0.3px;
    line-height: 1;
    margin-bottom: 2px;
}}
.banner-sub {{
    font-size: 7.5px;
    color: rgba(255,255,255,0.5);
}}
"""


# ---------------------------------------------------------------------------
# Company Cheat Sheet
# ---------------------------------------------------------------------------

def _src(d: dict, field: str, max_len=50) -> str:
    s = d.get(f"{field}_source") or d.get(f"{field}_tier") or ""
    return str(s)[:max_len]


def build_company_html(company: str, research: dict, model_values: dict,
                        intelligence: dict | None = None) -> str:
    today  = date.today().strftime("%B %d, %Y")
    basics = research.get("company_basics") or {}
    attio  = research.get("attio_insights") or {}
    gong   = research.get("gong_insights") or {}
    sources= research.get("source_summary") or {}
    comps  = research.get("comps_benchmarks") or {}
    vcfg   = load_vertical_config(research)

    # ── KPI strip ──
    revenue   = basics.get("annual_revenue")
    units     = basics.get("unit_count")
    employees = basics.get("employee_count")
    unit_label = vcfg.get("unit_label") or basics.get("unit_count_label", "Locations")

    kpi_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">▲</span>Key Metrics</div>
  <div class="kpi-strip">
    <div class="kpi-card">
      <div class="kpi-v">{fmt(revenue)}</div>
      <div class="kpi-l">Annual Revenue</div>
      <div class="kpi-src">{esc(_src(basics, 'annual_revenue'))}</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-v">{fmt_plain(units)}</div>
      <div class="kpi-l">{esc(unit_label)}</div>
      <div class="kpi-src">{esc(_src(basics, 'unit_count'))}</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-v">{fmt_plain(employees)}</div>
      <div class="kpi-l">Employees</div>
      <div class="kpi-src">{esc(_src(basics, 'employee_count'))}</div>
    </div>
  </div>
</div>"""

    # ── Company Profile (vertical-specific fields) ──
    profile_rows = []
    for fdef in vcfg.get("profile_fields", []):
        path = fdef.get("path", "")
        label = fdef.get("label", path)
        raw = _deep_get(research, path)
        if raw is None:
            continue
        fmt_mode = fdef.get("format", "text")
        if fmt_mode == "currency":
            display = fmt(raw)
        elif fmt_mode == "count":
            display = fmt_plain(raw)
        elif fmt_mode == "percent":
            display = fmt_plain(raw) if isinstance(raw, float) and raw < 1 else str(raw)
        else:
            if isinstance(raw, list):
                display = ", ".join(str(x) for x in raw)
            else:
                display = str(raw)
        profile_rows.append((label, display))

    # Geography: replace "States Operated" count with full states list if available
    geo = basics.get("geography") or {}
    if geo.get("states"):
        states_str = ", ".join(geo["states"])
        profile_rows = [(lbl, val) for lbl, val in profile_rows if lbl != "States Operated"]
        profile_rows.append(("States", states_str))

    profile_html = ""
    if profile_rows:
        rows_html = "".join(f"<tr><td>{esc(k)}</td><td>{esc(v)}</td></tr>" for k, v in profile_rows)
        profile_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">◈</span>Company Profile</div>
  <table class="data-table">{rows_html}</table>
</div>"""

    # ── Employee Breakdown ──
    breakdown_html = ""
    if vcfg.get("employee_breakdown") and basics.get("employee_breakdown"):
        bd = basics["employee_breakdown"]
        keys   = vcfg.get("employee_breakdown_keys") or list(bd.keys())
        labels = vcfg.get("employee_breakdown_labels") or [k.replace("_", " ").title() for k in keys]
        cells  = ""
        for k, lbl in zip(keys, labels):
            v = bd.get(k)
            if v is not None:
                cells += f'<div class="bk-cell"><div class="bk-v">{fmt_plain(v)}</div><div class="bk-l">{esc(lbl)}</div></div>'
        if cells:
            breakdown_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">⊞</span>Employee Breakdown</div>
  <div class="breakdown-grid">{cells}</div>
</div>"""

    # ── Industry Benchmarks ──
    bench_html = ""
    bench_rows = []
    for fdef in vcfg.get("benchmark_fields", []):
        path  = fdef.get("path", "")
        label = fdef.get("label", path)
        raw   = _deep_get(research, path)
        if raw is None:
            continue
        if isinstance(raw, dict):
            lo  = raw.get("low")
            mid = raw.get("mid")
            hi  = raw.get("high")
            bench_rows.append((label, lo, mid, hi))
        else:
            bench_rows.append((label, None, raw, None))
    if bench_rows:
        has_range = any(lo is not None or hi is not None for _, lo, _, hi in bench_rows)
        if has_range:
            hdr = '<tr class="assump-hdr"><th>Benchmark</th><th>Low</th><th>Mid</th><th>High</th></tr>'
            rows_html = hdr + "".join(
                f'<tr><td>{esc(lbl)}</td>'
                f'<td>{fmt_plain(lo) if lo is not None else "—"}</td>'
                f'<td class="mid">{fmt_plain(mid) if mid is not None else "—"}</td>'
                f'<td>{fmt_plain(hi) if hi is not None else "—"}</td></tr>'
                for lbl, lo, mid, hi in bench_rows
            )
            bench_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">≈</span>Industry Benchmarks</div>
  <table class="bench-table">{rows_html}</table>
</div>"""
        else:
            rows_html = "".join(
                f'<tr><td>{esc(lbl)}</td><td>{fmt_plain(mid) if mid is not None else "—"}</td></tr>'
                for lbl, _, mid, _ in bench_rows
            )
            bench_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">≈</span>Industry Benchmarks</div>
  <table class="data-table">{rows_html}</table>
</div>"""

    # ── Pain Points — check gong_insights first, then contextual_data ──
    pain_points = (
        gong.get("pain_points") or
        (research.get("contextual_data") or {}).get("pain_points") or
        []
    )
    pain_html = ""
    if pain_points:
        items = "".join(f"<li>{esc(str(p))}</li>" for p in pain_points[:8])
        pain_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">!</span>Key Pain Points</div>
  <ul class="bullets">{items}</ul>
</div>"""

    # ── Key Contacts (champions) ──
    champions = gong.get("champions") or []
    champs_html = ""
    if champions:
        rows_html = ""
        for ch in champions:
            if isinstance(ch, dict):
                rows_html += f"""<div class="champion-row">
  <div class="champ-name">{esc(ch.get('name',''))}</div>
  <div class="champ-title">{esc(ch.get('title',''))}</div>
  <div class="champ-note">{esc(ch.get('note',''))}</div>
</div>"""
        if rows_html:
            champs_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">✦</span>Key Contacts</div>
  <div style="background:{WHITE};border:1px solid {GRAY_LT};border-radius:6px;padding:4px 10px;margin-top:6px;">{rows_html}</div>
</div>"""

    # ── Tech Stack ──
    tech_stack = gong.get("tech_stack") or {}
    tech_html = ""
    if tech_stack and isinstance(tech_stack, dict):
        rows_html = "".join(
            f"<tr><td>{esc(k)}</td><td>{esc(v)}</td></tr>"
            for k, v in tech_stack.items()
        )
        tech_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">⚙</span>Tech Stack</div>
  <table class="tech-table">{rows_html}</table>
</div>"""

    # ── Verbatim Quotes ──
    quotes_html = ""
    raw_quotes = gong.get("verbatim_quotes") or []
    call_date  = gong.get("call_date") or ""
    interviewee= gong.get("interviewee") or gong.get("call_title") or ""
    if raw_quotes:
        blocks = ""
        for q in raw_quotes[:3]:
            text = q if isinstance(q, str) else (q.get("text") or q.get("quote") or "")
            src  = q.get("source", "") if isinstance(q, dict) else f"{interviewee} — {call_date}"
            if text:
                src_tag = f'<div class="quote-src">{esc(src)}</div>' if src else ""
                blocks += f'<div class="quote">"{esc(text)}"{src_tag}</div>'
        if blocks:
            quotes_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">❝</span>Verbatim Quotes</div>
  {blocks}
</div>"""

    # ── Research Sources ──
    pill_data = [
        ("Gong",  sources.get("gong_calls_found") or sources.get("gong_calls_transcribed") or 0),
        ("Attio", sources.get("attio_records", 0)),
        ("Slack", sources.get("slack_messages", 0)),
        ("SEC",   1 if sources.get("sec_filings") else 0),
        ("Web",   sources.get("web_operations_used", 0)),
    ]
    pills = "".join(
        f'<span class="pill {"pill-zero" if not n else ""}">'
        f'{esc(label)} <span class="pill-val">{n}</span></span>'
        for label, n in pill_data
    )
    source_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">◎</span>Research Sources</div>
  <div class="pills">{pills}</div>
</div>"""

    # ── Meeting Intelligence (Claude-generated) ──
    intel = intelligence or {}
    intel_html = ""
    if intel:
        parts = ""

        # Quick Take
        headlines = intel.get("headline_insights") or []
        if headlines:
            items = "".join(f"<li>{esc(h)}</li>" for h in headlines)
            parts += f"""<div class="intel-subsection">
  <div class="intel-sub-label">Quick Take</div>
  <div class="intel-hl"><ul class="bullets" style="margin-top:0;">{items}</ul></div>
</div>"""

        # Your Edge
        angle = intel.get("competitive_angle") or ""
        if angle:
            parts += f"""<div class="intel-subsection">
  <div class="intel-sub-label">Your Edge vs. Current Setup</div>
  <div class="intel-angle">{esc(angle)}</div>
</div>"""

        # Conversation Starters
        starters = intel.get("conversation_starters") or []
        if starters:
            items = "".join(f"<li>{esc(s)}</li>" for s in starters)
            parts += f"""<div class="intel-subsection">
  <div class="intel-sub-label">Lead With</div>
  <ol class="numbered-list">{items}</ol>
</div>"""

        # Deal Dynamics
        dynamics = intel.get("deal_dynamics") or []
        if dynamics:
            items = "".join(f"<li>{esc(d)}</li>" for d in dynamics)
            parts += f"""<div class="intel-subsection">
  <div class="intel-sub-label">Deal Dynamics</div>
  <ul class="bullets" style="margin-top:0;">{items}</ul>
</div>"""

        if parts:
            intel_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">⚡</span>Meeting Intelligence</div>
  {parts}
</div>"""

    # ── Objection Handling (from intelligence) ──
    obj_handling_html = ""
    obj_handling = intel.get("objection_handling") or {}
    raw_objections = gong.get("key_objections") or []
    if obj_handling and raw_objections:
        rows_html = ""
        for obj in raw_objections:
            # Try exact match first, then fuzzy
            response = obj_handling.get(obj)
            if not response:
                for k, v in obj_handling.items():
                    if k.lower()[:30] in obj.lower() or obj.lower()[:30] in k.lower():
                        response = v
                        break
            if response:
                rows_html += f"<tr><td>{esc(obj)}</td><td>{esc(response)}</td></tr>"
        if rows_html:
            obj_handling_html = f"""
<div class="section">
  <div class="section-label"><span class="s-icon">⚑</span>Key Objections + How to Respond</div>
  <table class="objection-table">
    <tr style="background:{NAVY};color:{WHITE};font-size:7.5px;font-weight:700;text-transform:uppercase;">
      <td style="padding:5px 9px;color:{WHITE};">Objection</td>
      <td style="padding:5px 9px;color:{WHITE};">Response</td>
    </tr>
    {rows_html}
  </table>
</div>"""

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>{CSS}</style>
</head>
<body>
<div class="banner">
  <div>
    <div class="banner-eyebrow">JOLLY.COM &nbsp;&middot;&nbsp; CONFIDENTIAL</div>
    <div class="banner-title">{esc(company)}</div>
    <div class="banner-sub">Company Cheat Sheet &nbsp;&middot;&nbsp; {today}</div>
  </div>
</div>
<div class="body">
{kpi_html}
{profile_html}
{breakdown_html}
{bench_html}
{intel_html}
{pain_html}
{champs_html}
{obj_handling_html}
{tech_html}
{quotes_html}
{source_html}
<div class="footer">
  <span>Jolly.com</span>
  <span>Confidential — Internal Use Only</span>
  <span>{today}</span>
</div>
</div>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Campaign Cheat Sheet
# ---------------------------------------------------------------------------

def build_campaign_html(company: str, research: dict, model_values: dict) -> str:
    today     = date.today().strftime("%B %d, %Y")
    campaigns = research.get("campaigns_selected") or []
    ci        = research.get("campaign_inputs") or {}

    # Build fuzzy lookup: for any research campaign slug, find the best matching model slug
    model_slugs = list({k[len("rops__"):] for k in model_values if k.startswith("rops__")} |
                       {k[len("ebitda__"):] for k in model_values if k.startswith("ebitda__")})

    def _mv(name: str, prefix: str):
        """Get a model value by campaign name, using fuzzy slug matching."""
        slug = slugify(name)
        exact = model_values.get(f"{prefix}{slug}")
        if exact is not None:
            return exact
        matched = _match_model_slug(slug, model_slugs)
        if matched:
            return model_values.get(f"{prefix}{matched}")
        return None

    normalised = []
    for c in campaigns:
        if isinstance(c, str):
            slug_c = slugify(c)
            inp    = ci.get(slug_c) or {}
            normalised.append({
                "name":            c,
                "priority":        inp.get("priority", "standard"),
                "evidence":        inp.get("evidence") or inp.get("evidence_source") or inp.get("rationale") or "",
                "client_interest": inp.get("client_interest") or inp.get("interest") or "",
                "rops":            _mv(c, "rops__") or inp.get("rops") or "",
                "savings":         _mv(c, "ebitda__") or inp.get("savings") or inp.get("ebitda_impact") or "",
                "include_summary": bool(
                    inp.get("include_in_summary_slide") or inp.get("include_summary_slide")
                    or inp.get("summary_slide") or inp.get("priority", "").upper() == "HIGH"
                ),
            })
        elif isinstance(c, dict):
            name_c = c.get("campaign_type") or c.get("name") or ""
            slug_c = slugify(name_c)
            inp    = ci.get(slug_c) or {}
            normalised.append({
                "name":            name_c,
                "priority":        c.get("priority") or "standard",
                "evidence":        c.get("evidence") or c.get("evidence_source") or inp.get("evidence") or inp.get("rationale") or "",
                "client_interest": c.get("client_interest") or c.get("interest") or inp.get("client_interest") or "",
                "rops":            _mv(name_c, "rops__") or c.get("rops") or "",
                "savings":         _mv(name_c, "ebitda__") or c.get("savings") or c.get("ebitda_impact") or "",
                "include_summary": bool(
                    c.get("include_summary_slide") or c.get("include_in_summary_slide")
                    or c.get("summary_slide")
                ),
            })

    if not normalised:
        cards_html = f'<p style="color:{MUTED};padding:16px 0;font-size:11px;">No campaign data. Run /deck-research first.</p>'
    else:
        cards = []
        for c in normalised:
            name     = c["name"]
            priority = c["priority"].lower()
            interest = c["client_interest"]
            evidence = c["evidence"]
            rops     = c["rops"]
            savings  = c["savings"]
            include  = c["include_summary"]
            is_high  = priority == "high"

            head_cls = "c-head high" if is_high else "c-head"
            bdg_cls  = "badge b-high" if is_high else "badge b-std"
            bdg_txt  = "High Priority" if is_high else priority.capitalize()

            # Client interest as a small header badge (not a row)
            interest_badge = ""
            if interest:
                interest_badge = f'<span class="badge b-interest">{esc(str(interest).capitalize())}</span>'

            # Quick stats row: ROPS | Est. Savings | Summary Slide
            rops_fmt = f"{rops:.0f}x" if isinstance(rops, (int, float)) else str(rops) if rops else "—"
            savings_fmt = fmt(savings) if savings else "—"
            include_fmt = "Yes" if include else "No"

            stats_html = f"""<div class="c-stats">
  <div class="c-stat-item">
    <div class="c-stat-v">{esc(rops_fmt)}</div>
    <div class="c-stat-l">ROPS</div>
  </div>
  <div class="c-stat-item">
    <div class="c-stat-v">{esc(savings_fmt)}</div>
    <div class="c-stat-l">Est. Savings</div>
  </div>
  <div class="c-stat-item">
    <div class="c-stat-v">{include_fmt}</div>
    <div class="c-stat-l">Summary Slide</div>
  </div>
</div>"""

            ev_block = f'<div class="evidence">"{esc(str(evidence))}"</div>' if evidence else ""

            # Assumptions block — exact match first, then fuzzy match using model slugs
            slug_name   = slugify(name)
            assump_rows = model_values.get(f"assumptions__{slug_name}")
            if not assump_rows:
                # Try fuzzy match against all assumptions__ keys
                assump_slugs = [k[len("assumptions__"):] for k in model_values if k.startswith("assumptions__")]
                matched_slug = _match_model_slug(slug_name, assump_slugs) if assump_slugs else None
                if matched_slug:
                    assump_rows = model_values.get(f"assumptions__{matched_slug}") or []
                else:
                    assump_rows = []
            assump_block = ""
            if assump_rows:
                # Check if rows are tuples of 4 (new format) or 2 (legacy)
                sample = assump_rows[0]
                if len(sample) == 4:
                    lbl_s, b_s, u_s, d_s = zip(*assump_rows)
                    has_scenarios = any(u or d for u, d in zip(u_s, d_s))
                else:
                    # Legacy 2-tuple: treat as base only
                    assump_rows = [(lbl, b, "", "") for lbl, b in assump_rows]
                    has_scenarios = False

                if has_scenarios:
                    hdr = '<tr class="assump-hdr"><th>Assumption</th><th>Base</th><th>Upside</th><th>Downside</th></tr>'
                    row_cells = "".join(
                        f'<tr><td>{esc(lbl)}</td>'
                        f'<td class="col-base">{esc(b)}</td>'
                        f'<td class="col-up">{esc(u) if u else "—"}</td>'
                        f'<td class="col-dn">{esc(d) if d else "—"}</td></tr>'
                        for lbl, b, u, d in assump_rows
                    )
                    tbl = f'<table class="assump-table">{hdr}{row_cells}</table>'
                    label_txt = "Model Assumptions — Base / Upside / Downside"
                else:
                    row_cells = "".join(
                        f'<tr><td>{esc(lbl)}</td><td class="col-base">{esc(b)}</td></tr>'
                        for lbl, b, *_ in assump_rows
                    )
                    tbl = f'<table class="assump-table">{row_cells}</table>'
                    label_txt = "Model Assumptions (Base)"

                assump_block = f"""<div class="assumptions">
  <div class="assumptions-label">{label_txt}</div>
  {tbl}
</div>"""

            cards.append(f"""
<div class="campaign">
  <div class="{head_cls}">
    <div class="c-title">{esc(name)}</div>
    <div class="c-badges">
      {interest_badge}
      <span class="{bdg_cls}">{bdg_txt}</span>
    </div>
  </div>
  <div class="c-body">
    {stats_html}
    {ev_block}
    {assump_block}
  </div>
</div>""")
        cards_html = "\n".join(cards)

    # Total Opportunity banner
    opp = _compute_total_opportunity(model_values, normalised)
    total_ebitda = opp["total_ebitda"]
    avg_rops     = opp["avg_rops"]
    high_count   = opp["high_count"]
    total_count  = opp["total_count"]

    opp_items = ""
    if total_ebitda:
        opp_items += f'<div class="opp-item"><div class="opp-v">{fmt(total_ebitda)}</div><div class="opp-l">Total EBITDA</div></div>'
    if avg_rops:
        opp_items += f'<div class="opp-item"><div class="opp-v">{avg_rops:.0f}x</div><div class="opp-l">Avg ROPS</div></div>'
    opp_items += f'<div class="opp-item"><div class="opp-v">{high_count}</div><div class="opp-l">High Priority</div></div>'
    opp_items += f'<div class="opp-item"><div class="opp-v">{total_count}</div><div class="opp-l">Campaigns</div></div>'
    opp_banner_html = f'<div class="opp-banner">{opp_items}</div>' if opp_items else ""
    summary_html = ""

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>{CSS}</style>
</head>
<body>
<div class="banner">
  <div>
    <div class="banner-eyebrow">JOLLY.COM &nbsp;&middot;&nbsp; CONFIDENTIAL</div>
    <div class="banner-title">{esc(company)}</div>
    <div class="banner-sub">Campaign Cheat Sheet &nbsp;&middot;&nbsp; {today}</div>
  </div>
</div>
<div class="body">
{opp_banner_html}
{cards_html}
<div class="footer">
  <span>Jolly.com</span>
  <span>Confidential — Internal Use Only</span>
  <span>{today}</span>
</div>
</div>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Combined Cheat Sheet (single document)
# ---------------------------------------------------------------------------

def _build_header_template(company: str) -> str:
    """Build an inline-styled HTML banner for Playwright's headerTemplate.

    This renders on every page of the PDF inside the top margin area.
    Playwright auto-fills .pageNumber and .totalPages classes.
    """
    today = date.today().strftime("%B %d, %Y")
    # Playwright wraps headerTemplate in a #header container with default padding.
    # Override it to zero so the banner sits flush at the page top.
    return (
        f'<style>#header {{ padding: 0 !important; margin: 0 !important; }}</style>'
        f'<div style="width:100%;height:60px;background:{BG};'
        f'-webkit-print-color-adjust:exact;print-color-adjust:exact;">'
        f'<div style="width:100%;height:53px;'
        f'background:linear-gradient(135deg,{NAVY} 0%,{NAVY_LT} 100%);'
        f'border-bottom:3px solid {GOLD};padding:7px 28px 6px;'
        f'display:flex;align-items:center;justify-content:space-between;'
        f'box-sizing:border-box;font-size:10px;'
        f'-webkit-print-color-adjust:exact;print-color-adjust:exact;'
        f"font-family:-apple-system,'Segoe UI','Helvetica Neue',Arial,sans-serif;\">"
        f'<div>'
        f'<div style="font-size:6.5px;font-weight:700;letter-spacing:1.8px;'
        f'text-transform:uppercase;color:{GOLD};margin-bottom:2px;">'
        f'JOLLY.COM &middot; CONFIDENTIAL</div>'
        f'<div style="font-size:14px;font-weight:800;color:#fff;'
        f'letter-spacing:-0.3px;line-height:1;margin-bottom:2px;">'
        f'{esc(company)}</div>'
        f'<div style="font-size:7.5px;color:rgba(255,255,255,0.5);">'
        f'Cheat Sheet &middot; {today}</div>'
        f'</div>'
        f'<div style="font-size:7.5px;color:rgba(255,255,255,0.4);">'
        f'Page <span class="pageNumber"></span> of <span class="totalPages"></span>'
        f'</div>'
        f'</div>'
        f'</div>'
    )


def _extract_body_div(full_html: str) -> str:
    """Extract the .body div (and its contents) from a full HTML document.

    Strips the DOCTYPE/html/head/body wrapper and the .banner div, returning
    just the ``<div class="body">...</div>`` block.
    """
    marker = '<div class="body">'
    start = full_html.index(marker)
    end = full_html.index('</body>')
    return full_html[start:end].strip()


def build_combined_html(company: str, research: dict, model_values: dict,
                        intelligence: dict | None = None) -> str:
    """Build a single HTML document containing both cheat sheets.

    The banner is NOT included in the body — it is injected on every page via
    Playwright's headerTemplate (see ``_build_header_template``).
    """
    company_full  = build_company_html(company, research, model_values, intelligence)
    campaign_full = build_campaign_html(company, research, model_values)

    company_section  = _extract_body_div(company_full)
    campaign_section = _extract_body_div(campaign_full)

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>{CSS}</style>
</head>
<body>
{company_section}
<div class="page-break"></div>
{campaign_section}
</body>
</html>"""


# ---------------------------------------------------------------------------
# PDF rendering
# ---------------------------------------------------------------------------

def render_pdf(html_content: str, pdf_path: str, header_html: str = "") -> bool:
    if not PLAYWRIGHT_AVAILABLE:
        print("  WARNING: Playwright not installed.")
        print("  Run: pip install playwright && playwright install chromium")
        return False
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            try:
                page = browser.new_page(viewport={"width": 816, "height": 1056})
                page.set_content(html_content, wait_until="networkidle")
                pdf_opts = {
                    "path": pdf_path,
                    "format": "Letter",
                    "print_background": True,
                }
                if header_html:
                    pdf_opts["display_header_footer"] = True
                    pdf_opts["header_template"] = header_html
                    pdf_opts["footer_template"] = "<span></span>"
                    pdf_opts["margin"] = {
                        "top": "60px", "bottom": "0",
                        "left": "0", "right": "0",
                    }
                else:
                    pdf_opts["margin"] = {
                        "top": "0", "bottom": "0",
                        "left": "0", "right": "0",
                    }
                page.pdf(**pdf_opts)
            finally:
                browser.close()
        return True
    except Exception as e:
        print(f"  ERROR rendering PDF: {e}")
        return False


# ---------------------------------------------------------------------------
# Save HTML templates
# ---------------------------------------------------------------------------

PLACEHOLDER_RESEARCH = {
    "company_name": "{{COMPANY_NAME}}",
    "research_date": "{{DATE}}",
    "industry": "Manufacturing / Food & Beverage",
    "company_basics": {
        "annual_revenue": 1_000_000_000,
        "annual_revenue_source": "Source",
        "unit_count": 500,
        "unit_count_source": "Source",
        "employee_count": 10_000,
        "employee_count_source": "Source",
        "geography": {"hq": "City, State", "states": ["State 1", "State 2"], "state_count": 5},
        "employee_breakdown": {
            "manufacturing_warehousing": 2500,
            "sales_distribution_delivery": 5000,
            "corporate_back_office": 2500,
        },
    },
    "attio_insights": {"founded": 1990},
    "gong_insights": {
        "pain_points": ["Example pain point 1", "Example pain point 2"],
        "champions": [{"name": "Jane Doe", "title": "CFO", "note": "Key champion"}],
        "key_objections": ["Example objection"],
        "tech_stack": {"System A": "Description of system A"},
        "verbatim_quotes": ["Example verbatim quote from sales call"],
    },
    "campaigns_selected": [
        {"name": "Campaign One", "priority": "high", "evidence": "Evidence from research",
         "client_interest": "Explicit", "include_summary_slide": True},
        {"name": "Campaign Two", "priority": "standard", "evidence": "Supporting evidence",
         "client_interest": "Implied", "include_summary_slide": False},
    ],
    "source_summary": {
        "gong_calls_found": 1, "attio_records": 1, "m365_emails": 0,
        "slack_messages": 0, "sec_filings": False, "web_operations_used": 4,
    },
    "comps_benchmarks": {
        "vertical": "manufacturing",
        "ebitda_margin_pct": {"low": 0.08, "mid": 0.14, "high": 0.22},
        "turnover_rate":     {"low": 0.20, "mid": 0.35, "high": 0.55},
        "gross_margin_pct":  {"low": 0.20, "mid": 0.30, "high": 0.45},
    },
    "campaign_inputs": {},
}


def save_templates():
    cfg = get_workspace_config()
    templates_root = cfg.get("templates_root", "Templates")
    tpl_dir = Path(templates_root) / "Cheat Sheets"
    tpl_dir.mkdir(parents=True, exist_ok=True)

    combined_tpl = build_combined_html("{{COMPANY_NAME}}", PLACEHOLDER_RESEARCH, {})
    (tpl_dir / "Cheat Sheet Template.html").write_text(combined_tpl, encoding="utf-8")
    print(f"  Template saved to {tpl_dir}/")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate cheat sheets for intro deck")
    parser.add_argument("--company", required=True)
    parser.add_argument("--client-path", default=None, help="Override base client folder path (for sub-brands nested under a parent)")
    args    = parser.parse_args()
    company = args.company

    print(f"\n=== cheatsheet_gen.py | {company} ===\n")

    if not OPENPYXL_AVAILABLE:
        print("WARNING: openpyxl not installed. Run: pip install openpyxl")

    cfg = get_workspace_config()
    client_root = cfg.get("client_root", "Clients")
    client_base = Path(args.client_path) if args.client_path else Path(client_root) / company

    try:
        research = find_research_json(company, base_path=str(client_base))
        print(f"  Research date: {research.get('research_date', 'unknown')}")
    except FileNotFoundError as e:
        print(f"  WARNING: {e}")
        research = {
            "company_name": company, "company_basics": {}, "attio_insights": {},
            "gong_insights": {}, "campaigns_selected": [], "campaign_inputs": {},
            "source_summary": {}, "comps_benchmarks": {}
        }

    try:
        model_path   = find_model(company, base_path=str(client_base))
        model_values = read_model_basics(model_path) if OPENPYXL_AVAILABLE else {}
        print(f"  Model: {model_path} ({len(model_values)} values)")
    except FileNotFoundError as e:
        print(f"  WARNING: {e}")
        model_values = {}

    out_dir = client_base / "4. Reports" / "Cheat Sheets"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Generate meeting intelligence (calls Claude if ANTHROPIC_API_KEY is set)
    reports_dir = str(out_dir)
    intelligence = _generate_meeting_intelligence(company, research, model_values,
                                                   cache_dir=reports_dir)

    # Build single combined HTML (company + campaign in one document)
    combined_html = build_combined_html(company, research, model_values, intelligence)
    header_html   = _build_header_template(company)

    pdf_path = out_dir / f"{company} Cheat Sheet.pdf"

    print("\nRendering Cheat Sheet...")
    ok = render_pdf(combined_html, str(pdf_path), header_html=header_html)
    print(f"  {'OK' if ok else 'FAILED'}: {pdf_path}")

    print("\nSaving templates...")
    save_templates()

    print("\nDone. Open with:")
    print(f'  start "" "{pdf_path}"')


if __name__ == "__main__":
    main()
