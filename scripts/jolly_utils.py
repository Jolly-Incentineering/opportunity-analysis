"""
Shared utilities for all Jolly agents.
Consolidates duplicated functions from populate scripts.
"""
import os
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

# ---------------------------------------------------------------------------
# Paths — resolved from JOLLY_WORKSPACE env var + workspace_config.json
# ---------------------------------------------------------------------------
BASE_DIR = Path(os.getenv("JOLLY_WORKSPACE", ".")).resolve()

# Read client_root and templates_root from workspace config if available
_ws_config_path = BASE_DIR / ".claude" / "data" / "workspace_config.json"
_ws_config = {}
if _ws_config_path.exists():
    try:
        _ws_config = json.loads(_ws_config_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        pass

CLIENTS_DIR = BASE_DIR / _ws_config.get("client_root", "Clients")
TEMPLATES_DIR = BASE_DIR / _ws_config.get("templates_root", "Templates")
CLAUDE_DIR = BASE_DIR / ".claude"

# Load .env if present
_env_path = CLAUDE_DIR / ".env"
if _env_path.exists():
    for _line in _env_path.read_text().splitlines():
        if "=" in _line and not _line.startswith("#"):
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

# Template paths
QSR_MODEL_TEMPLATE = TEMPLATES_DIR / "QSR" / "QSR Intro Model Template.xlsx"
QSR_PPT_TEMPLATE = TEMPLATES_DIR / "QSR" / "QSR Intro Template.pptx"
MFG_MODEL_TEMPLATE = TEMPLATES_DIR / "Manufacturing" / "Manufacturing Intro Template.xlsx"
MFG_PPT_TEMPLATE = TEMPLATES_DIR / "Manufacturing" / "Manufacturing Intro Template.pptx"
AUTO_MODEL_TEMPLATE = TEMPLATES_DIR / "Automotive Services"
TAXI_TEMPLATE = TEMPLATES_DIR / "Taxis"

# Comment defaults
COMMENT_AUTHOR = "Jolly Research"
COMMENT_WIDTH = 400
COMMENT_HEIGHT = 200

# Expected formula counts per template
FORMULA_COUNTS = {
    "qsr": {"Campaigns": 153, "Sensitivities": 86},
    "manufacturing": {"Campaigns": 366, "Sensitivities": 205},
}

# Accretion bounds
ACCRETION_BOUNDS = {
    "total_pct": (0.10, 0.15),
    "rops_per_campaign": (10, 30),
    "absolute_midsize": (3_000_000, 15_000_000),
    "absolute_small": (300_000, 3_000_000),
}

# Hours per employee per year by industry
HOURS_PER_YEAR = {
    "qsr": 1820,
    "manufacturing": 2080,
    "services": 2080,
}

# Hiring cost cap (QSR only)
QSR_HIRING_COST_CAP = 3500


# ---------------------------------------------------------------------------
# Excel comment helpers
# ---------------------------------------------------------------------------
def add_comment(ws, cell, text, author=COMMENT_AUTHOR, width=COMMENT_WIDTH,
                height=COMMENT_HEIGHT):
    """Add a formatted comment to an Excel cell."""
    c = Comment(text, author)
    c.width = width
    c.height = height
    ws[cell].comment = c


def set_cell(ws, ref, value, comment_text, author=COMMENT_AUTHOR,
             width=350, height=150):
    """Set cell value and attach comment in one call."""
    ws[ref] = value
    add_comment(ws, ref, comment_text, author=author, width=width,
                height=height)


def set_scenario_cells(ws, row, values, comment_text=None):
    """Set C/D/E cells for a row with Base/Upside/Downside values.

    *values* can be:
      - a single value (applied to all 3 scenarios)
      - a list/tuple of 3 values [base, upside, downside]
    """
    if not isinstance(values, (list, tuple)):
        values = [values, values, values]
    for col, val in zip(["C", "D", "E"], values):
        ws[f"{col}{row}"] = val
    if comment_text:
        # Add full comment to Base column
        add_comment(ws, f"C{row}", comment_text)
        # Add scenario-specific comments to Upside and Downside
        add_comment(ws, f"D{row}", f"UPSIDE SCENARIO\n\n{comment_text}")
        add_comment(ws, f"E{row}", f"DOWNSIDE SCENARIO\n\n{comment_text}")


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------
def load_workbook_safe(path, data_only=False):
    """Load an openpyxl workbook with clear error messaging."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Model not found: {path}")
    try:
        return load_workbook(path, data_only=data_only)
    except Exception as e:
        raise RuntimeError(f"Failed to load {path}: {e}") from e


def save_workbook_safe(wb, path):
    """Save workbook with error handling."""
    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            f"Cannot save {path} - file may be open in Excel. Close it first."
        )
    except Exception as e:
        raise RuntimeError(f"Failed to save {path}: {e}") from e


def count_formulas(ws):
    """Count formula cells (strings starting with '=') in a worksheet."""
    return sum(
        1 for row in ws.iter_rows() for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )


def verify_formula_counts(wb, template_type):
    """Verify formula counts match expected values for a template type.

    Returns (ok: bool, message: str).
    """
    expected = FORMULA_COUNTS.get(template_type)
    if not expected:
        return True, f"No formula counts defined for '{template_type}'"

    results = {}
    all_ok = True
    for sheet_name, expected_count in expected.items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
        if ws is None:
            results[sheet_name] = f"MISSING sheet"
            all_ok = False
            continue
        actual = count_formulas(ws)
        ok = actual == expected_count
        results[sheet_name] = f"{actual}/{expected_count} {'OK' if ok else 'MISMATCH'}"
        if not ok:
            all_ok = False

    msg = " | ".join(f"{k}: {v}" for k, v in results.items())
    return all_ok, msg


# ---------------------------------------------------------------------------
# Rounding helpers (Jolly illustrative model standards)
# ---------------------------------------------------------------------------
ROUNDING_RULES = {
    "revenue": -6,        # nearest $1M
    "revenue_500k": -5,   # nearest $500K for smaller cos
    "stores": 0,          # exact integer
    "orders": -1,         # nearest 10
    "employees_100": -2,  # nearest 100
    "employees_50": None, # nearest 50 (custom)
    "menu_price": 2,      # nearest $0.25 (custom)
    "margin_pct": 2,      # nearest 1% (0.01)
    "turnover_pct": 2,    # nearest 5% (custom)
    "hiring_cost_100": -2,
    "hiring_cost_500": None,
    "ebitda_per_hour": 2,  # nearest $0.25 (custom)
    "incentive": 2,
    "reduction_pct": 2,
}


def round_to_standard(value, field_type):
    """Round a value per Jolly rounding standards.

    Returns the rounded value.  For types needing custom rounding
    (e.g. nearest $0.25), handles specially.
    """
    if value is None:
        return None

    # Custom rounding for specific types
    if field_type == "menu_price" or field_type == "ebitda_per_hour":
        # Nearest $0.25
        return round(value * 4) / 4

    if field_type == "employees_50":
        return round(value / 50) * 50

    if field_type == "hiring_cost_500":
        return round(value / 500) * 500

    if field_type == "turnover_pct":
        # Nearest 5%
        return round(value * 20) / 20

    if field_type == "margin_pct":
        # Nearest 1%
        return round(value, 2)

    if field_type == "reduction_pct":
        # Nearest 2.5%
        return round(value * 40) / 40

    ndigits = ROUNDING_RULES.get(field_type)
    if ndigits is not None:
        return round(value, ndigits)

    return value


# ---------------------------------------------------------------------------
# EBITDA calculations
# ---------------------------------------------------------------------------
def calculate_ebitda_per_hour(revenue, employees, hours_per_year, ebitda_margin):
    """Calculate EBITDA per hour saved, rounded to nearest $0.25."""
    if employees == 0 or hours_per_year == 0:
        return 0
    raw = (revenue / employees / hours_per_year) * ebitda_margin
    return round_to_standard(raw, "ebitda_per_hour")


def calculate_orders_per_store_per_day(revenue, stores, aov):
    """Calculate orders per store per day from revenue / stores / 365 / AOV."""
    if stores == 0 or aov == 0:
        return 0
    return round(revenue / stores / 365 / aov)


# ---------------------------------------------------------------------------
# Template resolution
# ---------------------------------------------------------------------------
def get_template_paths(template_type):
    """Return (model_template, ppt_template) paths for a template type."""
    templates = {
        "qsr": (QSR_MODEL_TEMPLATE, QSR_PPT_TEMPLATE),
        "manufacturing": (MFG_MODEL_TEMPLATE, MFG_PPT_TEMPLATE),
    }
    result = templates.get(template_type.lower())
    if not result:
        raise ValueError(
            f"Unknown template type '{template_type}'. "
            f"Available: {list(templates.keys())}"
        )
    return result


def resolve_client_path(company_name, folder_override=None):
    """Resolve the client folder path.

    If *folder_override* is provided (e.g. 'Garnett Station/Firebirds'),
    use that relative to CLIENTS_DIR.  Otherwise use company_name directly.
    """
    rel = folder_override or company_name
    return CLIENTS_DIR / rel
