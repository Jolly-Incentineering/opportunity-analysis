"""
TEMPLATE SCANNER
Detects Excel template structure and compares against existing configs.

Finds matching configs for templates or creates new ones for custom templates.
"""

import json
from pathlib import Path
from typing import Dict, Optional, Tuple
from openpyxl import load_workbook
from difflib import SequenceMatcher


class TemplateScanner:
    """Scans Excel templates to extract structure and match against configs"""

    def __init__(self, templates_dir: Optional[Path] = None):
        """
        Initialize scanner with templates directory.

        Args:
            templates_dir: Path to .claude/agents/templates/
                          If None, uses default location
        """
        if templates_dir is None:
            templates_dir = Path(__file__).parent / "templates"

        self.templates_dir = templates_dir
        self.templates_dir.mkdir(exist_ok=True)

    def scan_template(self, excel_path: str) -> Dict:
        """
        Scan an Excel template and extract its structure.

        Returns:
            {
                "file_path": str,
                "template_type": "QSR" | "Manufacturing" | "Custom",
                "labels": {field: row_number, ...},
                "scenarios": ["C", "D", "E"],
                "structure_hash": str (for comparison)
            }
        """
        wb = load_workbook(excel_path)
        ws = wb["Inputs"]

        # Extract all labels from column B
        labels = {}
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=2, max_col=2):
            cell = row[0]
            if cell.value and isinstance(cell.value, str):
                labels[cell.value.strip()] = cell.row

        # Detect template type
        template_type = self._detect_template_type(labels)

        # Detect scenario columns
        scenarios = self._detect_scenarios(ws)

        # Create structure hash for comparison
        structure_hash = self._create_structure_hash(labels, template_type)

        result = {
            "file_path": str(excel_path),
            "template_type": template_type,
            "labels": labels,
            "scenarios": scenarios,
            "structure_hash": structure_hash,
        }

        wb.close()
        return result

    def find_matching_config(self, scanned_template: Dict) -> Tuple[Optional[str], float]:
        """
        Find a matching config for a scanned template.

        Args:
            scanned_template: Output from scan_template()

        Returns:
            (config_filename, similarity_score)
            - config_filename: Name of best matching config, or None if no match
            - similarity_score: 0-1 score (0.9+ is considered a match)
        """
        config_files = list(self.templates_dir.glob("*.json"))

        if not config_files:
            return None, 0.0

        best_match = None
        best_score = 0.0
        threshold = 0.85  # 85%+ similarity = match

        for config_file in config_files:
            try:
                with open(config_file, "r") as f:
                    config = json.load(f)

                # Compare labels
                score = self._compare_templates(
                    scanned_template["labels"],
                    config.get("labels", {})
                )

                if score > best_score:
                    best_score = score
                    if score >= threshold:
                        best_match = config_file.name

            except (json.JSONDecodeError, KeyError):
                continue

        return best_match, best_score

    def create_config_from_template(self, scanned_template: Dict, config_name: str) -> str:
        """
        Create a new config file from a scanned template.

        Args:
            scanned_template: Output from scan_template()
            config_name: Name for new config (e.g., "us_merchants_custom.json")

        Returns:
            Path to created config file
        """
        config = {
            "template_type": scanned_template["template_type"],
            "labels": scanned_template["labels"],
            "scenarios": scanned_template["scenarios"],
            "structure_hash": scanned_template["structure_hash"],
            "last_updated": self._get_timestamp(),
        }

        config_path = self.templates_dir / config_name

        with open(config_path, "w") as f:
            json.dump(config, f, indent=2)

        return str(config_path)

    def load_config(self, config_name: str) -> Dict:
        """Load a config file by name"""
        config_path = self.templates_dir / config_name

        if not config_path.exists():
            raise FileNotFoundError(f"Config not found: {config_name}")

        with open(config_path, "r") as f:
            return json.load(f)

    def _detect_template_type(self, labels: Dict[str, int]) -> str:
        """Detect template type based on labels present"""
        label_keys = {k.lower() for k in labels.keys()}

        # Retail indicators (REI standard)
        if any(x in label_keys for x in [
            "member sign-ups",
            "inventory accuracy",
            "employee referrals",
            "store count",
            "annual new members"
        ]):
            return "Retail"

        # Manufacturing indicators
        if any(x in label_keys for x in ["units produced", "defect rate", "trir"]):
            return "Manufacturing"

        # QSR indicators
        if any(x in label_keys for x in [
            "beverage contribution margin",
            "orders per store per day",
            "aov"
        ]):
            return "QSR"

        # Automotive indicators
        if any(x in label_keys for x in ["vehicle", "service", "repair"]):
            return "Automotive"

        return "Custom"

    def _detect_scenarios(self, ws) -> list:
        """Detect which columns contain scenario data (typically C, D, E)"""
        scenarios = []

        # Check columns C, D, E for common headers
        for col_letter in ["C", "D", "E"]:
            cell_value = ws[f"{col_letter}1"].value
            if cell_value and isinstance(cell_value, str):
                if any(x in cell_value.lower() for x in ["base", "upside", "downside"]):
                    scenarios.append(col_letter)

        # Default if not found
        if not scenarios:
            scenarios = ["C", "D", "E"]

        return scenarios

    def _create_structure_hash(self, labels: Dict[str, int], template_type: str) -> str:
        """Create a hash of template structure for comparison"""
        # Sort labels and combine with type for hash
        label_str = "|".join(sorted(labels.keys()))
        hash_input = f"{template_type}:{label_str}"

        # Simple hash (not cryptographic, just for comparison)
        import hashlib
        return hashlib.md5(hash_input.encode()).hexdigest()[:16]

    def _compare_templates(self, template_labels: Dict[str, int], config_labels: Dict[str, int]) -> float:
        """
        Compare two label sets using fuzzy matching.

        Returns: Similarity score 0-1
        """
        if not template_labels or not config_labels:
            return 0.0

        # Fuzzy match each template label to config labels
        matches = 0
        total = len(template_labels)

        for template_label in template_labels.keys():
            best_ratio = 0.0
            for config_label in config_labels.keys():
                ratio = SequenceMatcher(
                    None,
                    template_label.lower(),
                    config_label.lower()
                ).ratio()
                best_ratio = max(best_ratio, ratio)

            if best_ratio >= 0.8:  # 80%+ match on individual label
                matches += 1

        return matches / total if total > 0 else 0.0

    @staticmethod
    def _get_timestamp() -> str:
        """Get current timestamp"""
        from datetime import datetime
        return datetime.utcnow().isoformat()


# Example usage
if __name__ == "__main__":
    scanner = TemplateScanner()

    # Scan a template
    scanned = scanner.scan_template("Templates/QSR/QSR Intro Model Template.xlsx")
    print(f"Template type: {scanned['template_type']}")
    print(f"Labels found: {len(scanned['labels'])}")

    # Find matching config
    config_name, score = scanner.find_matching_config(scanned)
    print(f"Best match: {config_name} (score: {score:.2f})")

    if config_name:
        print(f"Using existing config: {config_name}")
    else:
        print("No match found, creating new config...")
        new_config = scanner.create_config_from_template(scanned, "qsr_custom.json")
        print(f"Created config: {new_config}")
