"""
Excel Editor Agent
==================
Unified agent for creating and editing Excel financial models.
Replaces: repopulate_qsr_models.py, populate_manufacturing_models.py,
          populate_garnett_station_models.py, populate_isg_value.py

Capabilities:
  - Populate QSR models (5 campaigns)
  - Populate Manufacturing models (10 campaigns)
  - Fix rounding violations
  - Adjust assumptions to hit accretion targets
  - Never overwrites formula cells
"""
import shutil
from pathlib import Path
from openpyxl import load_workbook

from jolly_utils import (
    BASE_DIR, CLIENTS_DIR,
    add_comment, set_cell, set_scenario_cells,
    load_workbook_safe, save_workbook_safe, verify_formula_counts,
    round_to_standard,
    calculate_ebitda_per_hour, calculate_orders_per_store_per_day,
    get_template_paths, resolve_client_path,
    QSR_HIRING_COST_CAP, HOURS_PER_YEAR, COMMENT_AUTHOR,
)


class ExcelEditor:
    """Creates and edits Jolly financial models."""

    def __init__(self, template_type="qsr"):
        self.template_type = template_type.lower()
        self.model_template, self.ppt_template = get_template_paths(
            self.template_type
        )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def create_model(self, company_name, folder=None, date_str=None):
        """Copy template to client folder and return the path.

        Returns Path to the new model file.
        """
        client_dir = resolve_client_path(company_name, folder)
        model_dir = client_dir / "1. Model"
        model_dir.mkdir(parents=True, exist_ok=True)

        if date_str is None:
            from datetime import date
            date_str = date.today().strftime("%Y.%m.%d")

        filename = f"{company_name} Intro Model ({date_str}).xlsx"
        dest = model_dir / filename
        shutil.copy2(self.model_template, dest)
        return dest

    def populate_qsr(self, model_path, data):
        """Populate a QSR model with company data.

        *data* dict keys:
          name, revenue, stores, aov, employees, bev_cost, food_cost,
          bev_margin, food_margin, turnover, hiring_cost, ebitda_margin,
          sources (dict of field->comment_text)
        """
        wb = load_workbook_safe(model_path)
        ws = wb["Inputs"]
        src = data.get("sources", {})

        # Company basics
        set_cell(ws, "E5", data["name"],
                 src.get("name", f"Company: {data['name']}"))
        set_cell(ws, "E6", data["revenue"], src.get("revenue", ""))
        set_cell(ws, "E7", data["stores"], src.get("stores", ""))

        # Orders per store per day (calculated)
        orders = calculate_orders_per_store_per_day(
            data["revenue"], data["stores"], data["aov"]
        )
        orders_comment = (
            f"Calculated: ${data['revenue']:,.0f} / {data['stores']} stores "
            f"/ 365 days / ${data['aov']:.2f} AOV = {orders}\n"
            f"{src.get('aov', '')}"
        )
        set_cell(ws, "E8", orders, orders_comment)
        set_cell(ws, "E9", data["employees"], src.get("employees", ""))

        # Campaign economics
        set_cell(ws, "E12", data["bev_cost"], src.get("bev_cost", ""))
        set_cell(ws, "E13", data["food_cost"], src.get("food_cost", ""))

        # Scenario assumptions (C/D/E columns)
        bm = data["bev_margin"]
        set_scenario_cells(ws, 30, bm, src.get("bev_margin", ""))

        fm = data["food_margin"]
        set_scenario_cells(ws, 34, fm, src.get("food_margin", ""))

        t = data["turnover"]
        set_scenario_cells(ws, 37, t, src.get("turnover", ""))

        # Hiring cost - C39 only, D39/E39 are formulas
        hiring = min(data["hiring_cost"], QSR_HIRING_COST_CAP)
        set_cell(ws, "C39", hiring, src.get("hiring_cost", ""))
        # Restore formulas
        ws["D39"] = "=$C39*(1+20%)"
        ws["E39"] = "=$C39*(1-20%)"

        # EBITDA per hour saved
        eh = calculate_ebitda_per_hour(
            data["revenue"], data["employees"],
            HOURS_PER_YEAR["qsr"], data["ebitda_margin"]
        )
        ebitda_comment = (
            f"Calculated: (${data['revenue']:,.0f} / {data['employees']} "
            f"/ {HOURS_PER_YEAR['qsr']} hrs) x {data['ebitda_margin']:.1%} "
            f"= ${eh:.2f}/hr\n{src.get('ebitda_margin', '')}"
        )
        set_scenario_cells(ws, 45, eh, ebitda_comment)

        save_workbook_safe(wb, model_path)
        ok, msg = verify_formula_counts(wb, "qsr")
        wb.close()
        return {"path": str(model_path), "formulas": msg, "ok": ok,
                "orders": orders, "ebitda_per_hour": eh}

    def populate_manufacturing(self, model_path, data):
        """Populate a Manufacturing model with company data.

        *data* dict keys: name, revenue, facilities, units_per_facility,
        employees, cost_defective, cost_per_unit, plus campaign-specific
        fields (return_rate, efficiency_gain, etc. as lists of 3).
        """
        wb = load_workbook_safe(model_path)
        ws = wb["Inputs"]

        # Company basics
        ws["E5"] = data["name"]
        add_comment(ws, "E5", data.get("name_src", data["name"]))
        ws["E6"] = data["revenue"]
        add_comment(ws, "E6", data.get("revenue_src", ""))
        ws["E7"] = data["facilities"]
        add_comment(ws, "E7", data.get("facilities_src", ""))
        ws["E8"] = data["units_per_facility"]
        add_comment(ws, "E8", data.get("units_src", ""))
        ws["E9"] = data["employees"]
        add_comment(ws, "E9", data.get("employees_src", ""))

        # Manufacturing economics
        ws["E14"] = data["cost_defective"]
        add_comment(ws, "E14", data.get("cost_defective_src", ""))
        ws["E15"] = data["cost_per_unit"]
        add_comment(ws, "E15", data.get("cost_per_unit_src", ""))

        # 10 Campaigns
        campaign_map = [
            # (row_start, fields)
            (35, [("return_rate", True), ("return_reduction", False),
                  ("cost_per_return", True)]),
            (40, [("efficiency_gain", False), ("gross_margin", True)]),
            (44, [("turnover_rate", True)]),
            # Campaign 3 hiring cost special: C46 only, D46/E46 formulas
            (52, [("ebitda_per_hour", True)]),
            (55, [("defect_rate", True), ("defect_reduction", False),
                  ("cost_per_defect", False)]),
            (60, [("quit_reduction", False), ("replacement_cost", True)]),
            (64, [("cross_trained", False), ("overtime_avoided", False),
                  ("overtime_premium", True)]),
            (69, [("safety_rate", True), ("safety_reduction", False),
                  ("cost_per_incident", True)]),
            (74, [("suggestions_per_emp", False),
                  ("savings_per_suggestion", False)]),
            (78, [("downtime_hours", True), ("downtime_reduction", False),
                  ("cost_per_downtime", True)]),
        ]

        for start_row, fields in campaign_map:
            for offset, (field_name, has_comment) in enumerate(fields):
                row = start_row + offset
                vals = data.get(field_name)
                if vals is not None:
                    set_scenario_cells(ws, row, vals)
                    if has_comment:
                        src_key = f"{field_name}_src"
                        src_text = data.get(src_key, "")
                        if src_text:
                            add_comment(ws, f"C{row}", src_text)

        # Campaign 3 hiring cost (C46 only)
        if "hiring_cost" in data:
            ws["C46"] = data["hiring_cost"]
            add_comment(ws, "C46", data.get("hiring_cost_src", ""))

        save_workbook_safe(wb, model_path)
        ok, msg = verify_formula_counts(wb, "manufacturing")
        wb.close()
        return {"path": str(model_path), "formulas": msg, "ok": ok}

    def fix_rounding(self, model_path, field_map):
        """Fix rounding violations in a model.

        *field_map*: dict of cell_ref -> (value, field_type)
        e.g. {"E6": (12247000, "revenue"), "C45": (4.33, "ebitda_per_hour")}
        """
        wb = load_workbook_safe(model_path)
        ws = wb["Inputs"]
        fixes = []

        for cell_ref, (current_val, field_type) in field_map.items():
            rounded = round_to_standard(current_val, field_type)
            if rounded != current_val:
                ws[cell_ref] = rounded
                fixes.append(
                    f"{cell_ref}: {current_val} -> {rounded} ({field_type})"
                )

        if fixes:
            save_workbook_safe(wb, model_path)
        wb.close()
        return {"fixes": fixes, "count": len(fixes)}

    def adjust_assumption(self, model_path, cell_ref, new_value,
                          comment_text=None):
        """Update a single assumption cell with optional comment."""
        wb = load_workbook_safe(model_path)
        ws = wb["Inputs"]

        # Safety: never overwrite formulas
        current = ws[cell_ref].value
        if isinstance(current, str) and current.startswith("="):
            wb.close()
            raise ValueError(
                f"Cell {cell_ref} contains formula '{current}' - "
                f"refusing to overwrite. Use a value cell instead."
            )

        ws[cell_ref] = new_value
        if comment_text:
            add_comment(ws, cell_ref, comment_text)

        save_workbook_safe(wb, model_path)
        wb.close()
        return {"cell": cell_ref, "value": new_value}

    def read_model_summary(self, model_path):
        """Read key metrics from a model for cross-file validation."""
        wb = load_workbook_safe(model_path, data_only=True)
        ws = wb["Inputs"]

        summary = {
            "company_name": ws["E5"].value,
            "revenue": ws["E6"].value,
            "stores_or_facilities": ws["E7"].value,
            "orders_or_units": ws["E8"].value,
            "employees": ws["E9"].value,
        }

        # Try to read campaign outputs
        if "Campaigns" in wb.sheetnames:
            cs = wb["Campaigns"]
            # Read total EBITDA accretion (typically last summary row)
            summary["campaigns_sheet_exists"] = True
        else:
            summary["campaigns_sheet_exists"] = False

        wb.close()
        return summary

    def batch_populate(self, companies_list, template_type=None):
        """Populate multiple companies sequentially.

        Each item in *companies_list* should have keys matching
        populate_qsr() or populate_manufacturing() signatures,
        plus 'folder' for client path resolution.
        """
        tt = template_type or self.template_type
        results = []

        for data in companies_list:
            name = data["name"]
            try:
                path = self.create_model(
                    name, folder=data.get("folder")
                )
                if tt == "qsr":
                    result = self.populate_qsr(path, data)
                elif tt == "manufacturing":
                    result = self.populate_manufacturing(path, data)
                else:
                    result = {"error": f"Unknown template type: {tt}"}
                result["company"] = name
                result["status"] = "success"
            except Exception as e:
                result = {"company": name, "status": "error", "error": str(e)}
            results.append(result)

        return results


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    import argparse, json as _json, sys

    parser = argparse.ArgumentParser(description="Excel editor for Jolly intro models")
    parser.add_argument("--file", required=True, help="Path to Excel model file")
    parser.add_argument("--action", required=True,
                        choices=["scan-formulas", "write-cells", "read-summary"],
                        help="Action to perform")
    parser.add_argument("--cells", help="JSON string of cell writes: [{cell, value, comment}, ...]")
    parser.add_argument("--template", default="qsr", help="Template type (default: qsr)")
    args = parser.parse_args()

    editor = ExcelEditor(args.template)

    if args.action == "scan-formulas":
        wb = load_workbook_safe(args.file, data_only=False)
        formula_cells = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells.append(f"{sheet_name}!{cell.coordinate}")
        wb.close()
        print(_json.dumps({"formula_cells": formula_cells, "count": len(formula_cells)}))

    elif args.action == "write-cells":
        if not args.cells:
            print("ERROR: --cells required for write-cells action", file=sys.stderr)
            sys.exit(1)
        writes = _json.loads(args.cells)
        wb = load_workbook_safe(args.file)
        ws = wb["Inputs"]
        written = 0
        for w in writes:
            ref = w["cell"]
            # Safety: never overwrite formulas
            current = ws[ref].value
            if isinstance(current, str) and current.startswith("="):
                print(f"SKIPPED {ref}: contains formula", file=sys.stderr)
                continue
            ws[ref] = w["value"]
            if w.get("comment"):
                add_comment(ws, ref, w["comment"])
            written += 1
        save_workbook_safe(wb, args.file)
        wb.close()
        print(_json.dumps({"written": written, "total": len(writes)}))

    elif args.action == "read-summary":
        result = editor.read_model_summary(args.file)
        print(_json.dumps(result, default=str))


if __name__ == "__main__":
    main()
